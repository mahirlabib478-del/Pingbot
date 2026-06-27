import os
import time
import threading
import requests
import datetime
import json
import io
import gzip
import uuid
import logging
import random
from flask import Flask
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ================== LOGGING ==================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================== CONFIG ==================
# Use environment variables for sensitive data
BOT_TOKEN = os.environ.get("BOT_TOKEN", "8808046131:AAG7g0k_hhvQV8cLRmh6ieKeuNBdBphfWkk")
ADMIN_CHAT_ID = os.environ.get("ADMIN_CHAT_ID", "2035024902")
CHANNEL_ID = os.environ.get("CHANNEL_ID", "-1003903695158")
BOT_USERNAME = os.environ.get("BOT_USERNAME", "Ping478bot")

if not BOT_TOKEN or not ADMIN_CHAT_ID:
    raise RuntimeError("BOT_TOKEN and ADMIN_CHAT_ID must be set")

# ================== FILE PATHS ==================
MOTHER_FILE = "mother_accounts.json"
COOLDOWN_FILE = "user_cooldowns.json"
SUBSCRIBERS_FILE = "subscribers.json"
USER_INFO_FILE = "user_info.json"
BALANCES_FILE = "balances.json"
GAME_BALANCES_FILE = "game_balances.json"
CONFIG_FILE = "config.json"
SUBMISSIONS_FILE = "submissions.json"
MOTHER_STOCK_FILE = "mother_stock.json"
REFERRALS_FILE = "referrals.json"
REFERRAL_BONUSES_FILE = "referral_bonuses.json"
LEADERBOARD_FILE = "leaderboard.json"
DEPOSITS_FILE = "deposits.json"
TRANSACTIONS_FILE = "transactions.json"
DUPLICATE_USERNAMES_FILE = "duplicate_usernames.json"
RPS_WINS_FILE = "rps_daily_wins.json"
USER_VERSIONS_FILE = "user_versions.json"
BACKUP_TRACKER_FILE = "backup_tracker.json"

app = Flask(__name__)

# ================== GLOBALS ==================
last_update_id = None
subscribed_users = set()
user_info = {}
mother_accounts = []
user_last_request = {}
maintenance_mode = False

user_balances = {}
game_balances = {}
submissions = []
mother_stock = []
config = {
    "price_cookies": 3.5,
    "price_2fa": 3.0,
    "mother_price": 5.0,
    "referral_level1": 5.0,
    "referral_level2": 1.0,
    "monthly_target": 5000.0,
    "target_bonus": 2.0,
    "lock_2fa": False,
    "lock_cookies": False,
    "bkash_number": "01XXXXXXXXX",
    "nagad_number": "01XXXXXXXXX",
    "channel_id": str(CHANNEL_ID) if CHANNEL_ID else "",
    "maintenance_mode": False,
    "bot_version": "1.0"
}
referrals = {}
referral_bonuses = {}
leaderboard = {}
withdraw_requests = []
deposit_requests = []
transactions = []
submitted_usernames = set()
rps_daily_wins = {}
user_versions = {}

# Session trackers
submission_sessions = {}
admin_approve_sessions = {}
admin_add_mother_session = {}
admin_add_mother_bulk_session = {}
withdraw_sessions = {}
deposit_sessions = {}
support_sessions = set()
broadcast_sessions = {}
rps_sessions = {}

# Track last activity of sessions for cleanup
session_activity = {}

data_lock = threading.RLock()
backup_lock = threading.Lock()
last_backup_message_id = None
last_backup_part_ids = []
last_backup_part_file_ids = []

last_morning_sent_date = None
last_evening_sent_date = None

# Thread-local session
_thread_local = threading.local()

def get_bot_session():
    if not hasattr(_thread_local, "session"):
        _thread_local.session = requests.Session()
        _thread_local.session.headers.update({"Connection": "keep-alive"})
    return _thread_local.session

# ================== DEBOUNCED SAVE ==================
save_scheduled = False
save_timer = None

def schedule_save():
    global save_scheduled, save_timer
    with data_lock:
        if not save_scheduled:
            save_scheduled = True
            if save_timer:
                save_timer.cancel()
            save_timer = threading.Timer(2.0, execute_save)
            save_timer.daemon = True
            save_timer.start()

def execute_save():
    global save_scheduled
    with data_lock:
        save_all()
        save_scheduled = False

# ================== FILE I/O ==================
def load_json(filename, default):
    try:
        with open(filename, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default

def save_json(filename, data):
    with data_lock:
        try:
            with open(filename, "w") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except IOError as e:
            logger.error(f"Error saving {filename}: {e}")

def load_all():
    global mother_accounts, user_last_request, subscribed_users, user_info
    global user_balances, game_balances, submissions, config, referrals, referral_bonuses, leaderboard
    global mother_stock, withdraw_requests, deposit_requests, maintenance_mode
    global transactions, submitted_usernames, rps_daily_wins, user_versions
    global last_backup_message_id, last_backup_part_ids, last_backup_part_file_ids

    mother_accounts = load_json(MOTHER_FILE, [])
    user_last_request = load_json(COOLDOWN_FILE, {})
    subscribed_users = set(load_json(SUBSCRIBERS_FILE, {"subscribed": []}).get("subscribed", []))
    user_info = load_json(USER_INFO_FILE, {})
    user_balances = load_json(BALANCES_FILE, {})
    game_balances = load_json(GAME_BALANCES_FILE, {})
    submissions = load_json(SUBMISSIONS_FILE, [])
    mother_stock = load_json(MOTHER_STOCK_FILE, [])
    referrals = load_json(REFERRALS_FILE, {})
    referral_bonuses = load_json(REFERRAL_BONUSES_FILE, {})
    leaderboard = load_json(LEADERBOARD_FILE, {})

    default_config = {
        "price_cookies": 3.5, "price_2fa": 3.0, "mother_price": 5.0,
        "referral_level1": 5.0, "referral_level2": 1.0, "monthly_target": 5000.0,
        "target_bonus": 2.0, "lock_2fa": False, "lock_cookies": False,
        "bkash_number": "01XXXXXXXXX", "nagad_number": "01XXXXXXXXX",
        "channel_id": str(CHANNEL_ID) if CHANNEL_ID else "",
        "maintenance_mode": False,
        "bot_version": "1.0"
    }
    loaded_config = load_json(CONFIG_FILE, default_config)
    for k in default_config:
        if k not in loaded_config:
            loaded_config[k] = default_config[k]
    config = loaded_config
    maintenance_mode = config.get("maintenance_mode", False)

    withdraw_requests = load_json("withdraw_requests.json", [])
    for w in withdraw_requests:
        if "method" not in w:
            w["method"] = "bkash"
            w["account_number"] = w.get("bkash", "")
    deposit_requests = load_json(DEPOSITS_FILE, [])
    transactions = load_json(TRANSACTIONS_FILE, [])
    submitted_usernames = set(load_json(DUPLICATE_USERNAMES_FILE, []))
    rps_daily_wins = load_json(RPS_WINS_FILE, {})
    user_versions = load_json(USER_VERSIONS_FILE, {})

    # Load backup tracker
    tracker = load_json(BACKUP_TRACKER_FILE, {"last_backup_message_id": None, "last_backup_part_ids": [], "last_backup_part_file_ids": []})
    last_backup_message_id = tracker.get("last_backup_message_id")
    last_backup_part_ids = tracker.get("last_backup_part_ids", [])
    last_backup_part_file_ids = tracker.get("last_backup_part_file_ids", [])

def save_all():
    save_json(MOTHER_FILE, mother_accounts)
    save_json(COOLDOWN_FILE, user_last_request)
    save_json(SUBSCRIBERS_FILE, {"subscribed": list(subscribed_users)})
    save_json(USER_INFO_FILE, user_info)
    save_json(BALANCES_FILE, user_balances)
    save_json(GAME_BALANCES_FILE, game_balances)
    save_json(SUBMISSIONS_FILE, submissions)
    save_json(MOTHER_STOCK_FILE, mother_stock)
    save_json(REFERRALS_FILE, referrals)
    save_json(REFERRAL_BONUSES_FILE, referral_bonuses)
    save_json(LEADERBOARD_FILE, leaderboard)
    save_json("withdraw_requests.json", withdraw_requests)
    save_json(DEPOSITS_FILE, deposit_requests)
    save_json(TRANSACTIONS_FILE, transactions)
    save_json(DUPLICATE_USERNAMES_FILE, list(submitted_usernames))
    save_json(RPS_WINS_FILE, rps_daily_wins)
    save_json(USER_VERSIONS_FILE, user_versions)
    save_json(CONFIG_FILE, config)
    save_json(BACKUP_TRACKER_FILE, {
        "last_backup_message_id": last_backup_message_id,
        "last_backup_part_ids": last_backup_part_ids,
        "last_backup_part_file_ids": last_backup_part_file_ids
    })
    # Removed save_data_to_channel() from here to avoid excessive backups

# ================== TELEGRAM HELPERS ==================
def send_telegram_message(text, chat_id, reply_markup=None, parse_mode=None):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": chat_id, "text": text}
    if reply_markup: payload["reply_markup"] = reply_markup
    if parse_mode: payload["parse_mode"] = parse_mode
    session = get_bot_session()
    for _ in range(3):
        try:
            resp = session.post(url, json=payload, timeout=10)
            if resp.status_code == 429:
                retry = resp.json().get("parameters", {}).get("retry_after", 2)
                time.sleep(retry)
                continue
            return resp
        except Exception as e:
            logger.error(f"Send error: {e}")
            time.sleep(1)
    return None

def send_telegram_document(file_bytes, filename, chat_id, caption=""):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    session = get_bot_session()
    try:
        files = {'document': (filename, file_bytes,
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        resp = session.post(url, data={"chat_id": chat_id, "caption": caption}, files=files, timeout=30)
        if resp.status_code == 200 and resp.json().get("ok"):
            return resp.json()
        return None
    except Exception as e:
        logger.error(f"Document send error: {e}")
        return None

def delete_message(chat_id, message_id):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage"
    session = get_bot_session()
    try:
        session.post(url, json={"chat_id": chat_id, "message_id": message_id}, timeout=10)
    except Exception as e:
        logger.error(f"Delete message error: {e}")

def broadcast_message(text):
    with data_lock:
        users = list(subscribed_users)
    to_remove = []
    for uid in users:
        try:
            resp = send_telegram_message(text, uid)
            if resp and resp.status_code == 403:
                to_remove.append(uid)
        except:
            to_remove.append(uid)
        time.sleep(0.05)
    if to_remove:
        with data_lock:
            for uid in to_remove:
                subscribed_users.discard(uid)
                user_info.pop(uid, None)
        save_all()

def broadcast_media(media_type, file_id, caption):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/send{media_type.capitalize()}"
    with data_lock:
        users = list(subscribed_users)
    to_remove = []
    session = get_bot_session()
    for uid in users:
        try:
            if media_type == "voice":
                payload = {"chat_id": uid, "voice": file_id}
            else:
                payload = {"chat_id": uid, media_type: file_id, "caption": caption}
            resp = session.post(url, json=payload, timeout=10)
            if resp.status_code == 403:
                to_remove.append(uid)
        except:
            pass
        time.sleep(0.05)
    if to_remove:
        with data_lock:
            for uid in to_remove:
                subscribed_users.discard(uid)
                user_info.pop(uid, None)
        save_all()

def answer_callback_query(callback_id, text=None):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery"
    payload = {"callback_query_id": callback_id}
    if text: payload["text"] = text
    session = get_bot_session()
    try:
        session.post(url, json=payload, timeout=5)
    except Exception as e:
        logger.error(f"Callback answer error: {e}")

# ================== CHANNEL BACKUP ==================
MAX_PART_SIZE = 45 * 1024 * 1024

def cleanup_old_channel_backup():
    global last_backup_message_id, last_backup_part_ids, last_backup_part_file_ids
    if not CHANNEL_ID: return
    try:
        session = get_bot_session()
        if last_backup_message_id:
            session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/unpinChatMessage",
                         json={"chat_id": CHANNEL_ID, "message_id": last_backup_message_id})
        for part_id in last_backup_part_ids:
            try:
                delete_message(CHANNEL_ID, part_id)
            except: pass
        if last_backup_message_id:
            try:
                delete_message(CHANNEL_ID, last_backup_message_id)
            except: pass
        last_backup_message_id = None
        last_backup_part_ids = []
        last_backup_part_file_ids = []
    except Exception as e:
        logger.error(f"Backup cleanup error: {e}")

def save_data_to_channel():
    global last_backup_message_id, last_backup_part_ids, last_backup_part_file_ids
    if not CHANNEL_ID: return
    with backup_lock:
        try:
            with data_lock:
                data = {
                    "subscribed_users": list(subscribed_users), "user_info": user_info,
                    "user_balances": user_balances, "game_balances": game_balances,
                    "submissions": submissions, "mother_stock": mother_stock,
                    "mother_accounts": mother_accounts, "config": config,
                    "referrals": referrals, "referral_bonuses": referral_bonuses,
                    "leaderboard": leaderboard, "withdraw_requests": withdraw_requests,
                    "deposit_requests": deposit_requests, "user_last_request": user_last_request,
                    "transactions": transactions, "submitted_usernames": list(submitted_usernames),
                    "rps_daily_wins": rps_daily_wins, "user_versions": user_versions,
                    "timestamp": datetime.datetime.now().isoformat()
                }
            json_bytes = json.dumps(data, indent=2, ensure_ascii=False).encode('utf-8')
            compressed = gzip.compress(json_bytes, compresslevel=6)

            new_backup_msg_id = None
            new_part_ids = []
            new_part_file_ids = []

            if len(compressed) <= MAX_PART_SIZE:
                filename = f"backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json.gz"
                session = get_bot_session()
                resp = session.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                    data={"chat_id": CHANNEL_ID},
                    files={"document": (filename, compressed, "application/gzip")},
                    timeout=60
                )
                if resp.status_code == 200 and resp.json().get("ok"):
                    new_backup_msg_id = resp.json()["result"]["message_id"]
                    new_part_ids = []
                    new_part_file_ids = []
                else:
                    return
            else:
                chunks = [compressed[i:i+MAX_PART_SIZE] for i in range(0, len(compressed), MAX_PART_SIZE)]
                total = len(chunks)
                timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                for idx, chunk in enumerate(chunks, 1):
                    part_filename = f"backup_{timestamp}_part{idx}of{total}.json.gz"
                    session = get_bot_session()
                    resp = session.post(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                        data={"chat_id": CHANNEL_ID, "caption": f"Part {idx}/{total}"},
                        files={"document": (part_filename, chunk, "application/gzip")},
                        timeout=60
                    )
                    if resp.status_code == 200 and resp.json().get("ok"):
                        new_part_ids.append(resp.json()["result"]["message_id"])
                        new_part_file_ids.append(resp.json()["result"]["document"]["file_id"])
                    else:
                        logger.error(f"Failed to send backup part {idx}/{total}")
                        return

                index_data = {
                    "backup_id": timestamp,
                    "parts": new_part_ids,
                    "total_parts": total,
                    "timestamp": timestamp,
                    "file_ids": new_part_file_ids
                }
                index_text = json.dumps(index_data)
                index_resp = send_telegram_message(index_text, CHANNEL_ID)
                if index_resp and index_resp.status_code == 200 and index_resp.json().get("ok"):
                    new_backup_msg_id = index_resp.json()["result"]["message_id"]
                else:
                    return

            if new_backup_msg_id:
                session = get_bot_session()
                session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/pinChatMessage", json={
                    "chat_id": CHANNEL_ID,
                    "message_id": new_backup_msg_id,
                    "disable_notification": True
                })

            cleanup_old_channel_backup()

            last_backup_message_id = new_backup_msg_id
            last_backup_part_ids = new_part_ids
            last_backup_part_file_ids = new_part_file_ids

            save_json(BACKUP_TRACKER_FILE, {
                "last_backup_message_id": last_backup_message_id,
                "last_backup_part_ids": last_backup_part_ids,
                "last_backup_part_file_ids": last_backup_part_file_ids
            })

        except Exception as e:
            logger.error(f"Channel backup error: {e}")

def auto_backup_loop():
    while True:
        time.sleep(86400)
        save_data_to_channel()

def restore_data_from_payload(compressed_bytes):
    decompressed = gzip.decompress(compressed_bytes)
    data = json.loads(decompressed.decode('utf-8'))
    required_keys = ["subscribed_users", "user_info", "user_balances", "config"]
    for key in required_keys:
        if key not in data:
            raise ValueError(f"Missing key '{key}' in backup data")
    return data

def apply_restored_data(data):
    global subscribed_users, user_info, user_balances, game_balances, submissions, mother_stock, mother_accounts
    global config, referrals, referral_bonuses, leaderboard, withdraw_requests, deposit_requests, user_last_request
    global transactions, submitted_usernames, rps_daily_wins, user_versions
    with data_lock:
        subscribed_users = set(data.get("subscribed_users", []))
        user_info = data.get("user_info", {})
        user_balances = data.get("user_balances", {})
        game_balances = data.get("game_balances", {})
        submissions = data.get("submissions", [])
        mother_stock = data.get("mother_stock", [])
        mother_accounts = data.get("mother_accounts", [])
        config = data.get("config", config)
        referrals = data.get("referrals", {})
        referral_bonuses = data.get("referral_bonuses", {})
        leaderboard = data.get("leaderboard", {})
        withdraw_requests = data.get("withdraw_requests", [])
        deposit_requests = data.get("deposit_requests", [])
        user_last_request = data.get("user_last_request", {})
        transactions = data.get("transactions", [])
        submitted_usernames = set(data.get("submitted_usernames", []))
        rps_daily_wins = data.get("rps_daily_wins", {})
        user_versions = data.get("user_versions", {})
    save_all()

def auto_restore_from_channel():
    global last_backup_message_id, last_backup_part_ids, last_backup_part_file_ids
    if not CHANNEL_ID: return
    try:
        session = get_bot_session()
        resp = session.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getChat?chat_id={CHANNEL_ID}", timeout=20).json()
        if not resp.get("ok"): return
        pinned = resp["result"].get("pinned_message")
        if not pinned: return

        if "document" in pinned:
            file_id = pinned["document"]["file_id"]
            file_info = session.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}", timeout=20).json()
            if not file_info.get("ok"): return
            file_path = file_info["result"]["file_path"]
            content = session.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}", timeout=60).content
            compressed = content
            last_backup_part_ids = []
            last_backup_part_file_ids = []
        elif "text" in pinned:
            index = json.loads(pinned["text"])
            file_ids = index.get("file_ids", [])
            if file_ids:
                combined = bytearray()
                for fid in file_ids:
                    file_info = session.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={fid}", timeout=20).json()
                    if not file_info.get("ok"): return
                    file_path = file_info["result"]["file_path"]
                    part_content = session.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}", timeout=60).content
                    combined.extend(part_content)
                compressed = bytes(combined)
                last_backup_part_ids = index.get("parts", [])
                last_backup_part_file_ids = file_ids
            else:
                part_ids = index.get("parts", [])
                if not part_ids: return
                combined = bytearray()
                for part_msg_id in part_ids:
                    msg_resp = session.get(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/getChatMessage?chat_id={CHANNEL_ID}&message_id={part_msg_id}",
                        timeout=20
                    ).json()
                    if not msg_resp.get("ok") or "document" not in msg_resp.get("result", {}):
                        logger.error(f"Missing part message {part_msg_id}")
                        return
                    file_id = msg_resp["result"]["document"]["file_id"]
                    file_info = session.get(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}",
                        timeout=20
                    ).json()
                    if not file_info.get("ok"): return
                    file_path = file_info["result"]["file_path"]
                    part_content = session.get(
                        f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}",
                        timeout=60
                    ).content
                    combined.extend(part_content)
                compressed = bytes(combined)
                last_backup_part_ids = part_ids
                last_backup_part_file_ids = []
        else:
            return

        data = restore_data_from_payload(compressed)
        apply_restored_data(data)
        with data_lock:
            last_backup_message_id = pinned["message_id"]
        logger.info("Data restored from channel backup successfully")
    except Exception as e:
        logger.error(f"Auto-restore error: {e}")

# ================== KEYBOARDS ==================
def get_main_keyboard(chat_id, chat_type="private"):
    if chat_type != "private":
        return {"keyboard": [["📊 লিডারবোর্ড", "👥 রেফারেল"]], "resize_keyboard": True}
    kb = [
        ["💼 একাউন্ট সাবমিট", "👤 প্রোফাইল"],
        ["👥 রেফারেল", "💰 ব্যালেন্স"],
        ["💳 ডিপোজিট", "💸 উইথড্র"],
        ["📊 লিডারবোর্ড", "🎁 ফ্রি মাদার একাউন্ট"],
        ["🛒 মাদার একাউন্ট কিনুন", "📞 সাপোর্ট"],
        ["🎮 RPS গেম"]
    ]
    if str(chat_id) == ADMIN_CHAT_ID:
        kb.append(["🛠️ অ্যাডমিন প্যানেল"])
    return {"keyboard": kb, "resize_keyboard": True}

def admin_panel_keyboard():
    return {
        "keyboard": [
            ["📊 সাবমিটেড ফাইল", "⚙️ মূল্য নির্ধারণ"],
            ["👥 রেফারেল বোনাস %", "🔒 সাবমিট লক"],
            ["📢 ব্রডকাস্ট", "➕ মাদার একাউন্ট যোগ"],
            ["📦 মাদার স্টক", "💰 মাদার মূল্য সেট"],
            ["📋 ইউজার লিস্ট", "✉️ ইউজারকে মেসেজ"],
            ["📁 ব্যাকআপ", "📥 রিস্টোর"],
            ["📥 ডিপোজিট রিকোয়েস্ট", "💳 উইথড্র রিকোয়েস্ট"],
            ["💳 বিকাশ নম্বর সেট", "💳 নগদ নম্বর সেট"],
            ["🔙 মূল মেনু"]
        ],
        "resize_keyboard": True
    }

# ================== EXCEL GENERATORS ==================
def generate_submission_excel(usernames, passwords, twofa_list, submitter_username):
    wb = Workbook()
    ws = wb.active
    ws.title = "Submission"
    ws.append(["Username", "Password", "2FA Key", "Submitter"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for i in range(len(usernames)):
        ws.append([usernames[i], passwords[i] if i < len(passwords) else "",
                   twofa_list[i] if i < len(twofa_list) else "", ""])
    if len(usernames) > 0:
        ws.cell(row=2, column=4, value=submitter_username)
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        ws.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def generate_mother_purchase_excel(accounts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mother Accounts"
    ws.append(["Username", "Password", "2FA Key"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for acc in accounts:
        ws.append([acc["username"], acc["password"], acc.get("fa_key", "")])
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except: pass
        ws.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

# ================== ENHANCED TRACKING ==================
def init_leaderboard_entry(user_id):
    uid = str(user_id)
    if uid not in leaderboard:
        leaderboard[uid] = {
            "total_submitted_2fa":0, "total_submitted_cookies":0,
            "total_ok_2fa":0, "total_ok_cookies":0,
            "total_income":0.0,
            "current_month_income":0.0,
            "last_month_income":0.0,
            "today_ok_2fa":0, "today_ok_cookies":0,
            "today_date": str(datetime.date.today()),
            "monthly_bonus_paid": False,
            "monthly_target": None,
            "current_month_key": f"{datetime.datetime.now().year}-{datetime.datetime.now().month}"
        }

def reset_daily_if_needed(user_id):
    uid = str(user_id)
    now = datetime.datetime.now()
    today_str = str(now.date())
    entry = leaderboard.get(uid)
    if not entry:
        init_leaderboard_entry(uid)
        entry = leaderboard[uid]
    if entry.get("today_date") != today_str:
        entry["today_ok_2fa"] = 0
        entry["today_ok_cookies"] = 0
        entry["today_date"] = today_str

def add_ok(user_id, acc_type, count, amount):
    uid = str(user_id)
    with data_lock:
        init_leaderboard_entry(uid)
        entry = leaderboard[uid]
        now = datetime.datetime.now()
        current_key = f"{now.year}-{now.month}"

        if entry.get("current_month_key") != current_key:
            last_income = entry.get("current_month_income", 0.0)
            entry["last_month_income"] = last_income
            target = entry.get("monthly_target")
            if target and last_income >= target and not entry.get("monthly_bonus_paid", False):
                bonus = last_income * config["target_bonus"] / 100.0
                user_balances[uid] = user_balances.get(uid, 0) + bonus
                entry["total_income"] += bonus
                # Schedule a message to be sent after releasing lock
                threading.Thread(target=send_telegram_message, args=(f"🎉 গত মাসের টার্গেট পূরণ! বোনাস {bonus} টাকা আপনার ব্যালেন্সে যোগ হয়েছে।", uid)).start()
                entry["monthly_bonus_paid"] = True
            entry["current_month_income"] = 0.0
            entry["monthly_bonus_paid"] = False
            entry["current_month_key"] = current_key

        reset_daily_if_needed(uid)

        entry[f"total_ok_{acc_type}"] += count
        entry["total_income"] += amount
        entry["current_month_income"] += amount
        entry[f"today_ok_{acc_type}"] += count
        schedule_save()

def get_target_progress(uid):
    uid = str(uid)
    entry = leaderboard.get(uid, {})
    target = entry.get("monthly_target")
    if not target:
        return None
    try:
        now = datetime.datetime.now()
        next_month = now.replace(day=28) + datetime.timedelta(days=4)
        last_day = next_month - datetime.timedelta(days=next_month.day)
        days_left = (last_day.date() - now.date()).days + 1
        if days_left < 1:
            days_left = 1

        price_2fa = config.get("price_2fa", 3.0)
        price_cookies = config.get("price_cookies", 3.5)

        current_income = entry.get("current_month_income", 0.0)
        remaining = max(0, target - current_income)
        daily_needed = remaining / days_left if days_left > 0 else remaining

        today_2fa = entry.get("today_ok_2fa", 0)
        today_cookies = entry.get("today_ok_cookies", 0)
        today_income = today_2fa * price_2fa + today_cookies * price_cookies

        return {
            "target": target,
            "current_income": current_income,
            "remaining": remaining,
            "days_left": days_left,
            "daily_income_needed": daily_needed,
            "daily_2fa_needed": daily_needed / price_2fa if price_2fa > 0 else 0,
            "daily_cookies_needed": daily_needed / price_cookies if price_cookies > 0 else 0,
            "today_ok_2fa": today_2fa,
            "today_ok_cookies": today_cookies,
            "today_income": today_income,
            "price_2fa": price_2fa,
            "price_cookies": price_cookies
        }
    except Exception as e:
        logger.error(f"get_target_progress error for {uid}: {e}")
        return None

# ================== PROFILE HELPERS ==================
def get_profile_text(uid):
    uid = str(uid)
    with data_lock:
        init_leaderboard_entry(uid)
        bal = user_balances.get(uid, 0)
        gb = game_balances.get(uid, 0)
        stats = leaderboard.get(uid, {})
        target_progress = get_target_progress(uid)

    msg = (
        f"👤 {user_info.get(uid, uid)}-এর প্রোফাইল\n\n"
        f"💰 মোট ইনকাম: {stats.get('total_income', 0)} টাকা\n"
        f"📊 মূল ব্যালেন্স: {bal} টাকা"
    )
    if gb > 0:
        msg += f"\n🎮 গেম ব্যালেন্স (মাদার কেনার জন্য): {gb} টাকা"

    msg += (
        f"\n📅 গত মাসের আয়: {stats.get('last_month_income', 0)} টাকা\n\n"
        f"📤 সাবমিট:\n"
        f"  🔐 2FA: {stats.get('total_submitted_2fa', 0)} টি (ওকে: {stats.get('total_ok_2fa', 0)})\n"
        f"  🍪 কুকিজ: {stats.get('total_submitted_cookies', 0)} টি (ওকে: {stats.get('total_ok_cookies', 0)})\n"
    )

    if target_progress:
        msg += (
            f"------------------\n"
            f"🎯 মাসিক টার্গেট: {target_progress['target']} টাকা\n"
            f"📈 চলতি মাসের আয়: {target_progress['current_income']} টাকা\n"
            f"⏳ বাকি: {target_progress['remaining']} টাকা ({target_progress['days_left']} দিন)\n"
            f"📆 আজকের আয়: {target_progress['today_income']} টাকা "
            f"(2FA: {target_progress['today_ok_2fa']} টি, কুকিজ: {target_progress['today_ok_cookies']} টি)\n"
            f"📌 আজকের প্রয়োজন: প্রায় {target_progress['daily_income_needed']:.1f} টাকা\n"
            f"   ↳ 2FA দিয়ে: {target_progress['daily_2fa_needed']:.1f} টি ({target_progress['price_2fa']} টাকা)\n"
            f"   ↳ কুকিজ দিয়ে: {target_progress['daily_cookies_needed']:.1f} টি ({target_progress['price_cookies']} টাকা)\n"
        )
    else:
        msg += "🎯 এখনো মাসিক টার্গেট সেট করেননি।\n"

    return msg

# ================== TRANSACTION HISTORY ==================
def record_transaction(user_id, type_, amount, description):
    uid = str(user_id)
    txn = {
        "id": f"txn_{uuid.uuid4().hex[:8]}",
        "user_id": uid,
        "type": type_,
        "amount": amount,
        "balance_after": user_balances.get(uid, 0),
        "timestamp": time.time(),
        "description": description
    }
    with data_lock:
        transactions.append(txn)
        if len(transactions) > 100000:
            transactions[:] = transactions[-100000:]
        schedule_save()

# ================== SUBMISSION SYSTEM ==================
def start_submission(chat_id, acc_type):
    submission_sessions[chat_id] = {"step": "username", "type": acc_type}
    session_activity[chat_id] = time.time()
    type_label = "🍪 কুকিজ একাউন্ট" if acc_type == "cookies" else "🔐 2FA একাউন্ট"
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    send_telegram_message(
        f"📋 {type_label} সাবমিট\n\nপ্রথমে আপনার **ইউজারনেম** লিস্ট দিন (প্রতি লাইনে একটি):\n\nবাতিল করতে নিচের বাটনে চাপুন বা /cancel লিখুন।",
        chat_id, reply_markup=cancel_kb
    )

def process_submission_step(chat_id, text, sender_username):
    if chat_id not in submission_sessions: return False
    session_activity[chat_id] = time.time()
    if text.strip().lower() in ["/cancel", "/start"]:
        submission_sessions.pop(chat_id, None)
        session_activity.pop(chat_id, None)
        send_telegram_message("❌ সাবমিট বাতিল করা হয়েছে।", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    session = submission_sessions[chat_id]
    step = session["step"]
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    if step == "username":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id, reply_markup=cancel_kb)
            return True
        duplicates = [u for u in lines if u in submitted_usernames]
        unique = [u for u in lines if u not in submitted_usernames]
        if duplicates:
            send_telegram_message(
                f"⚠️ {len(duplicates)} টি ইউজারনেম আগেই জমা হয়েছে। শুধু নতুন {len(unique)} টি নেওয়া হবে।",
                chat_id, reply_markup=cancel_kb
            )
        if not unique:
            send_telegram_message("❌ সমস্ত ইউজারনেম ডুপ্লিকেট! সাবমিট বাতিল।", chat_id)
            del submission_sessions[chat_id]
            session_activity.pop(chat_id, None)
            return True
        session["usernames"] = unique
        session["step"] = "password"
        send_telegram_message(f"🔑 এখন **পাসওয়ার্ড** লিস্ট দিন (প্রতি লাইনে একটি):\n\nআপনার ইউজারনেম সংখ্যা: {len(unique)}",
                             chat_id, reply_markup=cancel_kb)
        return True
    elif step == "password":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if len(lines) != len(session["usernames"]):
            send_telegram_message(f"❌ পাসওয়ার্ড সংখ্যা ({len(lines)}) ইউজারনেম সংখ্যার সাথে মেলে না।", chat_id, reply_markup=cancel_kb)
            return True
        session["passwords"] = lines
        session["step"] = "2fa"
        send_telegram_message("🔐 এখন **2FA কী** লিস্ট দিন (প্রতি লাইনে, ফাঁকা রাখা যাবে):", chat_id, reply_markup=cancel_kb)
        return True
    elif step == "2fa":
        raw_lines = text.splitlines()
        twofa_list = [l.strip() for l in raw_lines]
        while len(twofa_list) > len(session["usernames"]) and twofa_list[-1] == '':
            twofa_list.pop()
        if len(twofa_list) != len(session["usernames"]):
            send_telegram_message(f"❌ 2FA কী সংখ্যা ({len(twofa_list)}) ইউজারনেম সংখ্যার সাথে মেলে না।", chat_id, reply_markup=cancel_kb)
            return True
        with data_lock:
            for u in session["usernames"]:
                submitted_usernames.add(u)
            schedule_save()

        send_telegram_message("✅ আপনার ফাইল প্রক্রিয়াধীন। একটু অপেক্ষা করুন...", chat_id)
        threading.Thread(target=process_submission_heavy, args=(
            session["usernames"], session["passwords"], twofa_list,
            sender_username, chat_id, session["type"]
        ), daemon=True).start()

        del submission_sessions[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message("🔝", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    return False

def process_submission_heavy(usernames, passwords, twofa_list, sender_username, chat_id, acc_type):
    type_label = "কুকিজ" if acc_type == "cookies" else "2FA"
    filename = f"{sender_username}_{chat_id}_{len(usernames)}pcs_{type_label}.xlsx"
    excel_bytes = generate_submission_excel(usernames, passwords, twofa_list, sender_username)
    caption = (f"📥 {user_info.get(chat_id, 'Unknown')} (@{sender_username}) "
               f"একটি {type_label} একাউন্ট ফাইল সাবমিট করেছেন।\nমোট {len(usernames)} টি একাউন্ট।")
    resp = send_telegram_document(excel_bytes, filename, ADMIN_CHAT_ID, caption=caption)
    if resp:
        result = resp["result"]
        file_id = result["document"]["file_id"]
        msg_id = result["message_id"]
        sub_id = uuid.uuid4().hex[:10]
        with data_lock:
            submissions.append({
                "id": sub_id, "user_id": chat_id, "username": sender_username, "type": acc_type,
                "count": len(usernames), "file_id": file_id, "admin_message_id": msg_id,
                "timestamp": time.time(), "status": "pending",
                "usernames": list(usernames)
            })
            update_user_submission_stats(chat_id, acc_type, len(usernames))
            schedule_save()
        send_telegram_message("✅ আপনার ফাইল অ্যাডমিনের কাছে পাঠানো হয়েছে।", chat_id)
    else:
        send_telegram_message("⚠️ ফাইল পাঠাতে সমস্যা হয়েছে। পরে চেষ্টা করুন।", chat_id)

def update_user_submission_stats(user_id, acc_type, count):
    uid = str(user_id)
    with data_lock:
        init_leaderboard_entry(uid)
        leaderboard[uid][f"total_submitted_{acc_type}"] += count

# ================== ADMIN APPROVAL ==================
def admin_approve_start(sub_id):
    if ADMIN_CHAT_ID in admin_approve_sessions:
        send_telegram_message("⚠️ আগের অ্যাপ্রুভ প্রক্রিয়া শেষ করুন বা বাতিল করুন।", ADMIN_CHAT_ID)
        return
    admin_approve_sessions[ADMIN_CHAT_ID] = {"sub_id": sub_id, "step": "ok_count"}
    session_activity[ADMIN_CHAT_ID] = time.time()
    send_telegram_message("✅ কতটি আইডি ওকে হয়েছে? সংখ্যা লিখুন:", ADMIN_CHAT_ID,
                         reply_markup={"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]})

def process_admin_approve_step(chat_id, text):
    if chat_id != ADMIN_CHAT_ID or ADMIN_CHAT_ID not in admin_approve_sessions: return False
    session_activity[ADMIN_CHAT_ID] = time.time()
    if text.strip().lower() == "/cancel":
        del admin_approve_sessions[ADMIN_CHAT_ID]
        session_activity.pop(ADMIN_CHAT_ID, None)
        send_telegram_message("❌ বাতিল করা হয়েছে।", ADMIN_CHAT_ID)
        return True
    try:
        ok_count = int(text.strip())
        if ok_count < 0: raise ValueError
    except:
        send_telegram_message("⚠️ সঠিক সংখ্যা দিন।", ADMIN_CHAT_ID)
        return True
    sub_id = admin_approve_sessions[ADMIN_CHAT_ID]["sub_id"]
    with data_lock:
        for sub in submissions:
            if sub["id"] == sub_id and sub["status"] == "pending":
                user_id = str(sub["user_id"])
                init_leaderboard_entry(user_id)
                acc_type = sub["type"]
                total_sub = leaderboard[user_id].get(f"total_submitted_{acc_type}", 0)
                already_ok = leaderboard[user_id].get(f"total_ok_{acc_type}", 0)
                max_possible = min(sub["count"], total_sub - already_ok)
                if ok_count > max_possible:
                    send_telegram_message(
                        f"❌ সর্বোচ্চ {max_possible} টি আইডি ওকে করা যাবে। (সাবমিট: {sub['count']}, ইতিমধ্যে ওকে: {already_ok})",
                        ADMIN_CHAT_ID)
                    return True
                sub["status"] = "approved"
                sub["ok_count"] = ok_count
                price = config["price_2fa"] if acc_type == "2fa" else config["price_cookies"]
                amount = ok_count * price
                user_balances[user_id] = user_balances.get(user_id, 0) + amount
                add_ok(user_id, acc_type, ok_count, amount)
                # Distribute referral bonus (with messages sent after lock release)
                distribute_referral_bonus(user_id, amount)
                record_transaction(user_id, "submission_earning", amount, f"{acc_type.upper()} OK ({ok_count} pcs)")
                save_all()
                send_telegram_message(f"✅ সাবমিশন {sub_id} অ্যাপ্রুভ হয়েছে। {ok_count} আইডি ওকে, {amount} টাকা যোগ করা হয়েছে।", ADMIN_CHAT_ID)
                send_telegram_message(f"🎉 আপনার {ok_count} টি আইডি ওকে হয়েছে! {amount} টাকা আপনার ব্যালেন্সে যোগ হয়েছে।", user_id)
                break
        else:
            send_telegram_message("❌ সাবমিশন পাওয়া যায়নি বা ইতিমধ্যে প্রসেস করা হয়েছে।", ADMIN_CHAT_ID)
    del admin_approve_sessions[ADMIN_CHAT_ID]
    session_activity.pop(ADMIN_CHAT_ID, None)
    return True

def distribute_referral_bonus(user_id, amount):
    uid = str(user_id)
    bonus_messages = []
    with data_lock:
        if uid in referrals:
            referrer = str(referrals[uid])
            bonus1 = amount * config["referral_level1"] / 100.0
            if bonus1 > 0:
                user_balances[referrer] = user_balances.get(referrer, 0) + bonus1
                referral_bonuses[referrer] = referral_bonuses.get(referrer, 0) + bonus1
                record_transaction(referrer, "referral_bonus", bonus1, f"Referral Level 1 from {uid}")
                bonus_messages.append((referrer, f"🎁 রেফারেল বোনাস: {bonus1} টাকা ({config['referral_level1']}%) পেয়েছেন!"))
                update_leaderboard_income(referrer, bonus1)
            if referrer in referrals:
                grand_referrer = str(referrals[referrer])
                bonus2 = amount * config["referral_level2"] / 100.0
                if bonus2 > 0:
                    user_balances[grand_referrer] = user_balances.get(grand_referrer, 0) + bonus2
                    referral_bonuses[grand_referrer] = referral_bonuses.get(grand_referrer, 0) + bonus2
                    record_transaction(grand_referrer, "referral_bonus", bonus2, f"Referral Level 2 from {uid}")
                    bonus_messages.append((grand_referrer, f"🎁 রেফারেল বোনাস (লেভেল ২): {bonus2} টাকা ({config['referral_level2']}%) পেয়েছেন!"))
                    update_leaderboard_income(grand_referrer, bonus2)
    # Send messages after releasing lock
    for target, msg in bonus_messages:
        send_telegram_message(msg, target)

def update_leaderboard_income(user_id, amount):
    uid = str(user_id)
    init_leaderboard_entry(uid)
    leaderboard[uid]["total_income"] += amount

def reject_submission(sub_id):
    with data_lock:
        for sub in submissions:
            if sub["id"] == sub_id and sub["status"] == "pending":
                for username in sub.get("usernames", []):
                    submitted_usernames.discard(username)
                user_id = str(sub["user_id"])
                acc_type = sub["type"]
                if user_id in leaderboard:
                    leaderboard[user_id][f"total_submitted_{acc_type}"] -= sub["count"]
                sub["status"] = "rejected"
                save_all()
                return True
    return False

# ================== MOTHER ACCOUNT (FREE) ==================
def handle_get_free_mother(chat_id):
    now = time.time()
    last = user_last_request.get(str(chat_id), 0)
    if now - last < 600:
        wait = int((600 - (now - last))/60) + 1
        send_telegram_message(f"⏳ {wait} মিনিট পরে ফ্রি মাদার একাউন্ট নিতে পারবেন।", chat_id)
        return
    with data_lock:
        for acc in mother_accounts:
            if not acc.get("assigned_to"):
                acc["assigned_to"] = str(chat_id)
                acc["assigned_at"] = now
                user_last_request[str(chat_id)] = now
                record_transaction(chat_id, "free_mother", 0, f"Free mother account: {acc['username']}")
                schedule_save()
                msg = f"🎁 ফ্রি মাদার একাউন্ট:\n👤 ইউজারনেম: {acc['username']}\n🔑 পাসওয়ার্ড: {acc['password']}"
                if acc.get("fa_key"): msg += f"\n🔐 2FA: {acc['fa_key']}"
                send_telegram_message(msg, chat_id)
                return
    send_telegram_message("❌ এখন কোনো ফ্রি মাদার একাউন্ট নেই।", chat_id)

# ================== BUY MOTHER ACCOUNT ==================
def start_buy_mother(chat_id):
    with data_lock:
        available = [m for m in mother_stock if not m.get("sold")]
    if not available:
        send_telegram_message("❌ মাদার একাউন্ট স্টক খালি।", chat_id)
        return
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    send_telegram_message(
        f"🛒 মাদার একাউন্ট কিনুন\nপ্রতি পিস মূল্য: {config['mother_price']} টাকা\nবর্তমানে উপলব্ধ স্টক: {len(available)} টি\nআপনি কতটি কিনতে চান? (সংখ্যা লিখুন)\nবাতিল করতে /cancel",
        chat_id, reply_markup=cancel_kb
    )
    submission_sessions[chat_id] = {"step": "mother_qty", "type": "mother_buy"}
    session_activity[chat_id] = time.time()

def process_mother_buy_step(chat_id, text):
    if chat_id not in submission_sessions or submission_sessions[chat_id].get("type") != "mother_buy":
        return False
    session_activity[chat_id] = time.time()
    if text.strip().lower() == "/cancel":
        del submission_sessions[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message("❌ বাতিল করা হয়েছে।", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    try:
        qty = int(text.strip())
        if qty <= 0: raise ValueError
    except:
        send_telegram_message("⚠️ সঠিক সংখ্যা দিন।", chat_id)
        return True
    price = config["mother_price"]
    total = qty * price
    uid = str(chat_id)
    with data_lock:
        bal_main = user_balances.get(uid, 0)
        bal_game = game_balances.get(uid, 0)
        if bal_main + bal_game < total:
            send_telegram_message(f"❌ পর্যাপ্ত ব্যালেন্স নেই (মূল: {bal_main}, গেম: {bal_game})।", chat_id)
            del submission_sessions[chat_id]
            session_activity.pop(chat_id, None)
            return True
        available = [m for m in mother_stock if not m.get("sold")]
        if qty > len(available):
            send_telegram_message(f"❌ পর্যাপ্ত স্টক নেই।", chat_id)
            del submission_sessions[chat_id]
            session_activity.pop(chat_id, None)
            return True

        to_buy = []
        new_stock = []
        bought = 0
        for acc in mother_stock:
            if not acc.get("sold") and bought < qty:
                to_buy.append(acc)
                bought += 1
            else:
                new_stock.append(acc)
        mother_stock[:] = new_stock

        remaining = total
        game_used = 0
        main_used = 0
        if bal_game >= remaining:
            game_balances[uid] = bal_game - remaining
            game_used = remaining
            remaining = 0
        else:
            game_used = bal_game
            remaining -= bal_game
            game_balances[uid] = 0
            main_used = remaining
            user_balances[uid] = bal_main - main_used

        record_transaction(chat_id, "mother_purchase", -total,
                           f"Bought {qty} mother accounts (game={game_used}, main={main_used})")
        schedule_save()

    send_telegram_message("🔄 আপনার মাদার একাউন্ট প্রসেস হচ্ছে...", chat_id)
    threading.Thread(target=deliver_mother_purchase, args=(chat_id, to_buy, qty, total), daemon=True).start()
    del submission_sessions[chat_id]
    session_activity.pop(chat_id, None)
    send_telegram_message("🔝", chat_id, reply_markup=get_main_keyboard(chat_id))
    return True

def deliver_mother_purchase(chat_id, to_buy, qty, total):
    excel = generate_mother_purchase_excel(to_buy)
    if send_telegram_document(excel, f"mother_{chat_id}_{int(time.time())}.xlsx", chat_id,
                              caption=f"{qty} টি মাদার একাউন্ট কেনা হয়েছে। মোট মূল্য: {total} টাকা"):
        with data_lock:
            for acc in to_buy:
                acc["sold"] = True
            schedule_save()
        send_telegram_message(f"✅ {qty} টি মাদার একাউন্ট কেনা সফল।", chat_id)
    else:
        with data_lock:
            mother_stock.extend(to_buy)
            user_balances[str(chat_id)] = user_balances.get(str(chat_id), 0) + total
            schedule_save()
        send_telegram_message("⚠️ ডেলিভারি ব্যর্থ। টাকা ফেরত দেওয়া হয়েছে।", chat_id)

# ================== ADMIN ADD MOTHER STOCK ==================
def start_add_mother_stock(chat_id):
    admin_add_mother_session[chat_id] = {"step": "username"}
    session_activity[chat_id] = time.time()
    send_telegram_message("➕ মাদার একাউন্ট যোগ করুন\nপ্রথমে ইউজারনেম লিস্ট দিন:", chat_id,
                         reply_markup={"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]})

def process_add_mother_step(chat_id, text):
    if chat_id not in admin_add_mother_session: return False
    session_activity[chat_id] = time.time()
    if text.strip().lower() == "/cancel":
        del admin_add_mother_session[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message("❌ বাতিল করা হয়েছে।", chat_id, reply_markup=admin_panel_keyboard())
        return True
    session = admin_add_mother_session[chat_id]
    step = session["step"]
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    if step == "username":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id, reply_markup=cancel_kb)
            return True
        session["usernames"] = lines
        session["step"] = "password"
        send_telegram_message(f"🔑 এখন পাসওয়ার্ড লিস্ট দিন ({len(lines)} টি):", chat_id, reply_markup=cancel_kb)
        return True
    elif step == "password":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if len(lines) != len(session["usernames"]):
            send_telegram_message("❌ পাসওয়ার্ড সংখ্যা মেলেনি।", chat_id, reply_markup=cancel_kb)
            return True
        session["passwords"] = lines
        session["step"] = "2fa"
        send_telegram_message("🔐 2FA কী লিস্ট দিন:", chat_id, reply_markup=cancel_kb)
        return True
    elif step == "2fa":
        raw_lines = text.splitlines()
        twofa_list = [l.strip() for l in raw_lines]
        while len(twofa_list) > len(session["usernames"]) and twofa_list[-1] == '':
            twofa_list.pop()
        if len(twofa_list) != len(session["usernames"]):
            send_telegram_message("❌ 2FA কী সংখ্যা মেলেনি।", chat_id, reply_markup=cancel_kb)
            return True
        with data_lock:
            for i in range(len(session["usernames"])):
                mother_stock.append({"username": session["usernames"][i], "password": session["passwords"][i],
                                     "fa_key": twofa_list[i], "sold": False})
            schedule_save()
        send_telegram_message(f"✅ {len(session['usernames'])} টি মাদার একাউন্ট যোগ হয়েছে।", chat_id, reply_markup=admin_panel_keyboard())
        del admin_add_mother_session[chat_id]
        session_activity.pop(chat_id, None)
        return True
    return False

def start_add_mother_bulk(chat_id):
    admin_add_mother_bulk_session[chat_id] = {"step": "username"}
    session_activity[chat_id] = time.time()
    send_telegram_message("➕ ফ্রি মাদার একাউন্ট বাল্ক যোগ\n\nপ্রথমে **ইউজারনেম** লিস্ট দিন (প্রতি লাইনে একটি):", chat_id,
                         reply_markup={"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]})

def process_add_mother_bulk_step(chat_id, text):
    if chat_id not in admin_add_mother_bulk_session: return False
    session_activity[chat_id] = time.time()
    if text.strip().lower() == "/cancel":
        del admin_add_mother_bulk_session[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message("❌ বাতিল করা হয়েছে।", chat_id, reply_markup=admin_panel_keyboard())
        return True
    session = admin_add_mother_bulk_session[chat_id]
    step = session["step"]
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    if step == "username":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            send_telegram_message("⚠️ কমপক্ষে একটি ইউজারনেম দিন।", chat_id, reply_markup=cancel_kb)
            return True
        session["usernames"] = lines
        session["step"] = "password"
        send_telegram_message(f"🔑 এখন **পাসওয়ার্ড** লিস্ট দিন ({len(lines)} টি):", chat_id, reply_markup=cancel_kb)
        return True
    elif step == "password":
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if len(lines) != len(session["usernames"]):
            send_telegram_message("❌ পাসওয়ার্ড সংখ্যা মেলেনি।", chat_id, reply_markup=cancel_kb)
            return True
        session["passwords"] = lines
        session["step"] = "2fa"
        send_telegram_message("🔐 এখন **2FA কী** লিস্ট দিন (প্রতি লাইনে, ফাঁকা রাখা যাবে):", chat_id, reply_markup=cancel_kb)
        return True
    elif step == "2fa":
        raw_lines = text.splitlines()
        twofa_list = [l.strip() for l in raw_lines]
        while len(twofa_list) > len(session["usernames"]) and twofa_list[-1] == '':
            twofa_list.pop()
        if len(twofa_list) != len(session["usernames"]):
            send_telegram_message("❌ 2FA কী সংখ্যা মেলেনি।", chat_id, reply_markup=cancel_kb)
            return True
        with data_lock:
            for i in range(len(session["usernames"])):
                mother_accounts.append({
                    "username": session["usernames"][i],
                    "password": session["passwords"][i],
                    "fa_key": twofa_list[i],
                    "assigned_to": None,
                    "assigned_at": None
                })
            schedule_save()
        send_telegram_message(f"✅ {len(session['usernames'])} টি ফ্রি মাদার একাউন্ট যোগ করা হয়েছে।", chat_id, reply_markup=admin_panel_keyboard())
        del admin_add_mother_bulk_session[chat_id]
        session_activity.pop(chat_id, None)
        return True
    return False

# ================== MOTHER STOCK DETAIL & DELETE ==================
def show_mother_stock_detail(chat_id, page=0, message_id=None):
    with data_lock:
        available = [(i, acc) for i, acc in enumerate(mother_stock) if not acc.get("sold")]
    if not available:
        send_telegram_message("📦 মাদার স্টক খালি।", chat_id)
        return

    items_per_page = 10
    total_pages = (len(available) + items_per_page - 1) // items_per_page
    page = max(0, min(page, total_pages-1)) if total_pages > 0 else 0
    start = page * items_per_page
    end = start + items_per_page
    page_items = available[start:end]

    lines = [f"📦 **পেইড মাদার স্টক (পৃষ্ঠা {page+1}/{total_pages}):**\n"]
    keyboard_rows = []
    for pos, (orig_idx, acc) in enumerate(page_items, start=start+1):
        line = f"{pos}. 👤 {acc['username']} | 🔑 {acc['password']} | 🔐 {acc.get('fa_key','')}"
        lines.append(line)
        keyboard_rows.append([{"text": f"🗑️ #{pos}", "callback_data": f"delmotherstock_{orig_idx}"}])

    nav_buttons = []
    if page > 0:
        nav_buttons.append({"text": "⬅️ পূর্ববর্তী", "callback_data": f"motherstock_page_{page-1}"})
    if page < total_pages - 1:
        nav_buttons.append({"text": "➡️ পরবর্তী", "callback_data": f"motherstock_page_{page+1}"})
    if nav_buttons:
        keyboard_rows.append(nav_buttons)

    keyboard_rows.append([{"text": "🔙 বন্ধ করুন", "callback_data": "close_motherstock"}])

    if message_id:
        session = get_bot_session()
        session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText", json={
            "chat_id": chat_id,
            "message_id": message_id,
            "text": "\n".join(lines),
            "reply_markup": {"inline_keyboard": keyboard_rows}
        })
    else:
        send_telegram_message("\n".join(lines), chat_id, reply_markup={"inline_keyboard": keyboard_rows})

def show_mother_stock_detail_refresh(chat_id, message_id, page=None):
    with data_lock:
        available = [(i, acc) for i, acc in enumerate(mother_stock) if not acc.get("sold")]
    if not available:
        try:
            session = get_bot_session()
            session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText", json={
                "chat_id": chat_id, "message_id": message_id,
                "text": "📦 মাদার স্টক খালি।",
            })
        except: pass
        return

    items_per_page = 10
    total_pages = (len(available) + items_per_page - 1) // items_per_page
    if page is None or page >= total_pages:
        page = total_pages - 1
    show_mother_stock_detail(chat_id, page=page, message_id=message_id)

def show_free_mother_list(chat_id, page=0, message_id=None):
    with data_lock:
        if not mother_accounts:
            send_telegram_message("🎁 ফ্রি মাদার একাউন্ট তালিকা খালি।", chat_id)
            return
        items_per_page = 10
        total_pages = (len(mother_accounts) + items_per_page - 1) // items_per_page
        page = max(0, min(page, total_pages-1)) if total_pages > 0 else 0
        start = page * items_per_page
        end = start + items_per_page
        page_items = list(enumerate(mother_accounts, 1))[start:end]

        lines = [f"🎁 ফ্রি মাদার একাউন্ট তালিকা (পৃষ্ঠা {page+1}/{total_pages}):\n"]
        keyboard_rows = []
        for idx, acc in page_items:
            assigned = "কেহ না" if not acc.get("assigned_to") else acc["assigned_to"]
            lines.append(f"{idx}. 👤 {acc['username']} | 🔑 {acc['password']} | 🔐 {acc.get('fa_key','')} | বরাদ্দ: {assigned}")
            keyboard_rows.append([{"text": f"🗑️ #{idx}", "callback_data": f"delfreemother_{idx-1}"}])

        nav_buttons = []
        if page > 0:
            nav_buttons.append({"text": "⬅️ পূর্ববর্তী", "callback_data": f"freemother_page_{page-1}"})
        if page < total_pages - 1:
            nav_buttons.append({"text": "➡️ পরবর্তী", "callback_data": f"freemother_page_{page+1}"})
        if nav_buttons:
            keyboard_rows.append(nav_buttons)

        keyboard_rows.append([{"text": "🔙 বন্ধ করুন", "callback_data": "close_freemotherlist"}])

    if message_id:
        session = get_bot_session()
        session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText", json={
            "chat_id": chat_id, "message_id": message_id,
            "text": "\n".join(lines),
            "reply_markup": {"inline_keyboard": keyboard_rows}
        })
    else:
        send_telegram_message("\n".join(lines), chat_id, reply_markup={"inline_keyboard": keyboard_rows})

def show_free_mother_list_refresh(chat_id, message_id, page=None):
    with data_lock:
        if not mother_accounts:
            try:
                session = get_bot_session()
                session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText", json={
                    "chat_id": chat_id, "message_id": message_id,
                    "text": "🎁 ফ্রি মাদার একাউন্ট তালিকা খালি।",
                })
            except: pass
            return
        items_per_page = 10
        total_pages = (len(mother_accounts) + items_per_page - 1) // items_per_page
        if page is None or page >= total_pages:
            page = total_pages - 1
        show_free_mother_list(chat_id, page=page, message_id=message_id)

# ================== PROFILE & LEADERBOARD ==================
def show_profile(chat_id):
    try:
        msg = get_profile_text(chat_id)
        inline_kb = {"inline_keyboard": [
            [{"text": "🎯 টার্গেট সেট করুন", "callback_data": "set_target"}],
            [{"text": "📜 ইতিহাস", "callback_data": "show_history"}]
        ]}
        send_telegram_message(msg, chat_id, reply_markup=inline_kb, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"show_profile error for {chat_id}: {e}")
        send_telegram_message("⚠️ প্রোফাইল দেখাতে সমস্যা হয়েছে। পরে চেষ্টা করুন।", chat_id)

def show_leaderboard(chat_id):
    with data_lock:
        sorted_users = sorted(leaderboard.items(), key=lambda x: x[1].get("total_income", 0), reverse=True)[:10]
    if not sorted_users:
        send_telegram_message("এখনো কোনো ইনকাম রেকর্ড নেই।", chat_id)
        return
    msg = "🏆 লিডারবোর্ড\n\n" + "\n".join(f"{i}. {user_info.get(uid, uid)} - {data.get('total_income',0)} টাকা"
                                          for i, (uid, data) in enumerate(sorted_users, 1))
    send_telegram_message(msg, chat_id)

# ================== WITHDRAW (dual payment) ==================
def start_withdraw(chat_id):
    withdraw_sessions[chat_id] = {"step": "method"}
    session_activity[chat_id] = time.time()
    kb = {
        "inline_keyboard": [
            [{"text": "💸 বিকাশ", "callback_data": "withmethod_bkash"}],
            [{"text": "💸 নগদ", "callback_data": "withmethod_nagad"}],
            [{"text": "❌ বাতিল", "callback_data": "cancel_session"}]
        ]
    }
    send_telegram_message("💸 উইথড্রের মাধ্যম বাছাই করুন:", chat_id, reply_markup=kb)

def process_withdraw_step(chat_id, text):
    if chat_id not in withdraw_sessions: return False
    session_activity[chat_id] = time.time()
    if text.strip().lower() in ["/cancel", "/start"]:
        del withdraw_sessions[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message("❌ উইথড্র বাতিল করা হয়েছে।", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    session = withdraw_sessions[chat_id]
    step = session["step"]
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    if step == "method":
        return False
    if step == "amount":
        try:
            amount = float(text.strip())
            if amount <= 0: raise ValueError
        except:
            send_telegram_message("⚠️ সঠিক সংখ্যা দিন।", chat_id, reply_markup=cancel_kb)
            return True
        if amount > user_balances.get(str(chat_id), 0):
            send_telegram_message("❌ অপর্যাপ্ত ব্যালেন্স।", chat_id)
            del withdraw_sessions[chat_id]
            session_activity.pop(chat_id, None)
            return True
        session["amount"] = amount
        session["step"] = "account"
        method_label = session["method"].upper()
        send_telegram_message(f"📞 আপনার {method_label} অ্যাকাউন্ট নম্বর দিন:", chat_id, reply_markup=cancel_kb)
        return True
    elif step == "account":
        account = text.strip()
        if not account:
            send_telegram_message("⚠️ নম্বর খালি রাখা যাবে না।", chat_id, reply_markup=cancel_kb)
            return True
        w_id = uuid.uuid4().hex[:10]
        amount = session["amount"]
        # Deduct balance immediately
        uid = str(chat_id)
        with data_lock:
            user_balances[uid] = user_balances.get(uid, 0) - amount
            withdraw_requests.append({
                "id": w_id, "user_id": uid, "amount": amount,
                "method": session["method"], "account_number": account,
                "status": "pending", "time": time.time()
            })
            record_transaction(chat_id, "withdraw", -amount, f"Withdraw request {w_id}")
            save_all()
        del withdraw_sessions[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message(f"✅ {amount} টাকা উইথড্র রিকোয়েস্ট জমা হয়েছে। ব্যালেন্স থেকে কেটে রাখা হয়েছে।", chat_id)
        send_telegram_message(
            f"💳 নতুন উইথড্র রিকোয়েস্ট\nআইডি: {w_id}\nইউজার: {user_info.get(uid, uid)}\n"
            f"পরিমাণ: {amount}\nমাধ্যম: {session['method'].upper()}\nঅ্যাকাউন্ট: {account}\n"
            f"/approvewithdraw {w_id} or /rejectwithdraw {w_id}",
            ADMIN_CHAT_ID)
        send_telegram_message("🔝", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    return False

# ================== DEPOSIT SYSTEM (dual payment) ==================
def start_deposit(chat_id):
    deposit_sessions[chat_id] = {"step": "method"}
    session_activity[chat_id] = time.time()
    kb = {
        "inline_keyboard": [
            [{"text": "💰 বিকাশ", "callback_data": "depmethod_bkash"}],
            [{"text": "💰 নগদ", "callback_data": "depmethod_nagad"}],
            [{"text": "❌ বাতিল", "callback_data": "cancel_session"}]
        ]
    }
    send_telegram_message("💳 ডিপোজিটের মাধ্যম বাছাই করুন:", chat_id, reply_markup=kb)

def process_deposit_step(chat_id, text):
    if chat_id not in deposit_sessions: return False
    session_activity[chat_id] = time.time()
    if text.strip().lower() in ["/cancel", "/start"]:
        del deposit_sessions[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message("❌ ডিপোজিট বাতিল করা হয়েছে।", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    session = deposit_sessions[chat_id]
    step = session["step"]
    cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
    if step == "method":
        return False
    if step == "amount":
        try:
            amount = float(text.strip())
            if amount <= 0: raise ValueError
        except:
            send_telegram_message("⚠️ সঠিক টাকার পরিমাণ দিন (শুধু সংখ্যা)।", chat_id, reply_markup=cancel_kb)
            return True
        session["amount"] = amount
        session["step"] = "trxid"
        method = session["method"]
        number = config.get(f"{method}_number", "সেট করা হয়নি")
        send_telegram_message(
            f"🔢 এখন আপনার {method.upper()} ট্রানজেকশন আইডি (TrxID) দিন:\n(আপনার {method.upper()} নম্বর থেকে {number} নম্বরে টাকা পাঠিয়ে TrxID দিন)",
            chat_id, reply_markup=cancel_kb)
        return True
    elif step == "trxid":
        trxid = text.strip()
        if not trxid:
            send_telegram_message("⚠️ ট্রানজেকশন আইডি খালি রাখা যাবে না।", chat_id, reply_markup=cancel_kb)
            return True
        amount = session["amount"]
        dep_id = uuid.uuid4().hex[:10]
        dep_req = {
            "id": dep_id, "user_id": str(chat_id), "amount": amount,
            "trxid": trxid, "method": session["method"],
            "status": "pending", "time": time.time()
        }
        with data_lock:
            deposit_requests.append(dep_req)
            save_all()
        del deposit_sessions[chat_id]
        session_activity.pop(chat_id, None)
        send_telegram_message(
            f"✅ আপনার {amount} টাকার ডিপোজিট রিকোয়েস্ট ({session['method'].upper()}) জমা হয়েছে।\nট্রানজেকশন আইডি: {trxid}\nঅ্যাডমিন অনুমোদন করলেই ব্যালেন্স যোগ হবে।",
            chat_id)
        admin_msg = (f"📥 নতুন ডিপোজিট রিকোয়েস্ট\nআইডি: {dep_id}\nইউজার: {user_info.get(str(chat_id), chat_id)} ({chat_id})\n"
                     f"পরিমাণ: {amount} টাকা\nমাধ্যম: {session['method'].upper()}\nট্রানজেকশন আইডি: {trxid}\n"
                     f"অনুমোদন: /approvedeposit {dep_id}\nবাতিল: /rejectdeposit {dep_id}")
        send_telegram_message(admin_msg, ADMIN_CHAT_ID)
        send_telegram_message("🔝", chat_id, reply_markup=get_main_keyboard(chat_id))
        return True
    return False

# ================== SUPPORT ==================
def start_support(chat_id):
    support_sessions.add(chat_id)
    session_activity[chat_id] = time.time()
    send_telegram_message("📞 আপনার মেসেজ, ছবি, ফাইল বা ভয়েস পাঠান। অ্যাডমিন সরাসরি দেখতে পাবেন।\nবাতিল করতে নিচের বাটনে চাপুন।",
                         chat_id, reply_markup={"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]})

def forward_support_message(chat_id, msg):
    session_activity[chat_id] = time.time()
    sender = user_info.get(str(chat_id), chat_id)
    session = get_bot_session()
    if "text" in msg:
        send_telegram_message(f"📩 সাপোর্ট মেসেজ\nইউজার: {sender} ({chat_id})\n\n{msg['text']}", ADMIN_CHAT_ID)
    elif "photo" in msg:
        session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
                     json={"chat_id": ADMIN_CHAT_ID, "photo": msg["photo"][-1]["file_id"],
                           "caption": f"📩 সাপোর্ট ছবি\nইউজার: {sender} ({chat_id})\n{msg.get('caption','')}"})
    elif "document" in msg:
        session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                     json={"chat_id": ADMIN_CHAT_ID, "document": msg["document"]["file_id"],
                           "caption": f"📩 সাপোর্ট ফাইল\nইউজার: {sender} ({chat_id})\n{msg.get('caption','')}"})
    elif "voice" in msg:
        session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendVoice",
                     json={"chat_id": ADMIN_CHAT_ID, "voice": msg["voice"]["file_id"],
                           "caption": f"📩 সাপোর্ট ভয়েস\nইউজার: {sender} ({chat_id})"})
    cancel_kb = {"inline_keyboard": [[{"text": "❌ সাপোর্ট বন্ধ করুন", "callback_data": "cancel_session"}]]}
    send_telegram_message("✅ আপনার মেসেজ পাঠানো হয়েছে।\nসাপোর্ট থেকে বের হতে নিচের বাটনে চাপুন অথবা /cancel লিখুন।", chat_id, reply_markup=cancel_kb)

# ================== ADMIN BROADCAST ==================
def admin_broadcast_prompt(chat_id):
    kb = {
        "inline_keyboard": [
            [{"text": "📝 টেক্সট", "callback_data": "bc_text"}],
            [{"text": "🖼️ ছবি", "callback_data": "bc_photo"}],
            [{"text": "📄 ফাইল", "callback_data": "bc_document"}],
            [{"text": "🎤 ভয়েস", "callback_data": "bc_voice"}],
            [{"text": "❌ বাতিল", "callback_data": "cancel_broadcast"}]
        ]
    }
    send_telegram_message("📢 কী ধরনের ব্রডকাস্ট করবেন?", chat_id, reply_markup=kb)

# ================== RPS GAME ==================
def start_rps(chat_id):
    today = str(datetime.date.today())
    with data_lock:
        entry = rps_daily_wins.setdefault(str(chat_id), {"date": today, "wins": 0})
        if entry.get("date") != today:
            entry["date"] = today
            entry["wins"] = 0
        schedule_save()
    rps_sessions[chat_id] = True
    session_activity[chat_id] = time.time()
    kb = {
        "inline_keyboard": [
            [{"text": "🪨 Rock", "callback_data": "rps_rock"},
             {"text": "📄 Paper", "callback_data": "rps_paper"},
             {"text": "✂️ Scissors", "callback_data": "rps_scissors"}],
            [{"text": "❌ বাতিল", "callback_data": "cancel_session"}]
        ]
    }
    send_telegram_message("🎮 Rock Paper Scissors!\nআপনার পছন্দ বাছাই করুন:", chat_id, reply_markup=kb)

def process_rps_callback(chat_id, user_choice):
    if chat_id not in rps_sessions: return
    del rps_sessions[chat_id]
    session_activity.pop(chat_id, None)
    choices = ["rock", "paper", "scissors"]
    bot_choice = random.choice(choices)
    win_map = {"rock": "scissors", "paper": "rock", "scissors": "paper"}
    if user_choice == bot_choice:
        result = "draw"
    elif win_map[user_choice] == bot_choice:
        result = "win"
    else:
        result = "lose"

    emoji = {"rock": "🪨", "paper": "📄", "scissors": "✂️"}
    msg = f"আপনি: {emoji[user_choice]}, বট: {emoji[bot_choice]}\n\n"
    if result == "win":
        msg += "🎉 আপনি জিতেছেন!"
        with data_lock:
            today = str(datetime.date.today())
            entry = rps_daily_wins.setdefault(str(chat_id), {"date": today, "wins": 0})
            if entry.get("date") != today:
                entry["date"] = today
                entry["wins"] = 0
            wins_today = entry["wins"]
            if wins_today < 3:
                entry["wins"] = wins_today + 1
                schedule_save()
                mother = None
                for acc in mother_accounts:
                    if not acc.get("assigned_to"):
                        mother = acc
                        break
                if mother:
                    mother["assigned_to"] = str(chat_id)
                    mother["assigned_at"] = time.time()
                    record_transaction(chat_id, "rps_reward", 0, f"RPS Game: Free mother account {mother['username']}")
                    reward_text = f"🎁 ফ্রি মাদার একাউন্ট: {mother['username']} / {mother['password']}"
                    if mother.get("fa_key"):
                        reward_text += f"\n🔐 2FA: {mother['fa_key']}"
                else:
                    game_balances[str(chat_id)] = game_balances.get(str(chat_id), 0) + 5
                    record_transaction(chat_id, "rps_reward", 5, "RPS Game: 5 TK (game balance)")
                    reward_text = "💰 গেম ব্যালেন্সে ৫ টাকা যোগ হয়েছে (শুধু মাদার একাউন্ট কেনার জন্য ব্যবহার করা যাবে)।"
                msg += f"\n{reward_text}\n(আজকের পুরস্কার {entry['wins']}/3 ব্যবহৃত)"
            else:
                msg += "\nদুঃখিত, আজ পুরস্কার শেষ। তারপরও ভালো খেলেছেন!"
    elif result == "draw":
        msg += "🤝 ড্র!"
    else:
        msg += "😞 আপনি হেরেছেন। আবার চেষ্টা করুন!"

    send_telegram_message(msg, chat_id, reply_markup=get_main_keyboard(chat_id))

# ================== INLINE MODE ==================
def handle_inline_query(inline_query):
    query_id = inline_query["id"]
    query_text = inline_query.get("query", "").strip()
    results = []
    if not query_text:
        results.append({
            "type": "article",
            "id": "1",
            "title": "একাউন্ট সাবমিট করুন",
            "input_message_content": {"message_text": "💼 একাউন্ট সাবমিট করতে চাপুন:"},
            "reply_markup": {"inline_keyboard": [[{"text": "সাবমিট করুন", "url": f"https://t.me/{BOT_USERNAME}"}]]}
        })
        results.append({
            "type": "article",
            "id": "2",
            "title": "RPS গেম খেলুন",
            "input_message_content": {"message_text": "🎮 RPS গেম খেলতে চাপুন:"},
            "reply_markup": {"inline_keyboard": [[{"text": "খেলুন", "url": f"https://t.me/{BOT_USERNAME}?start=rps"}]]}
        })
        results.append({
            "type": "article",
            "id": "3",
            "title": "প্রোফাইল দেখুন",
            "input_message_content": {"message_text": "👤 প্রোফাইল দেখতে চাপুন:"},
            "reply_markup": {"inline_keyboard": [[{"text": "প্রোফাইল", "url": f"https://t.me/{BOT_USERNAME}"}]]}
        })
    else:
        results.append({
            "type": "article",
            "id": "search",
            "title": f"Search: {query_text}",
            "input_message_content": {"message_text": f"🔍 {query_text} এর জন্য বটে যান:"},
            "reply_markup": {"inline_keyboard": [[{"text": "বট খুলুন", "url": f"https://t.me/{BOT_USERNAME}"}]]}
        })
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/answerInlineQuery"
    payload = {"inline_query_id": query_id, "results": json.dumps(results), "cache_time": 0}
    try:
        session = get_bot_session()
        session.post(url, json=payload, timeout=10)
    except Exception as e:
        logger.error(f"Inline answer error: {e}")

# ================== COMMAND HANDLER ==================
def handle_commands(chat_id, text, chat_type="private", msg=None):
    global maintenance_mode, config, user_balances, game_balances, user_info, subscribed_users
    parts = text.split()
    cmd = parts[0].lower()
    uid = str(chat_id)

    # --- User commands ---
    if cmd == "/start":
        if chat_type == "private":
            with data_lock:
                subscribed_users.add(uid)
                schedule_save()
            support_sessions.discard(uid)
            if len(parts) > 1:
                arg = parts[1]
                if arg.startswith("ref_") and arg[4:].isdigit() and arg[4:] != uid and uid not in referrals:
                    ref_id = arg[4:]
                    referrals[uid] = ref_id
                    schedule_save()
                    send_telegram_message(f"🎉 আপনি {user_info.get(ref_id, ref_id)}-এর রেফারেলে যুক্ত হয়েছেন!", uid)
                elif arg == "rps":
                    start_rps(uid)
                    return
        send_telegram_message("✨ স্বাগতম! নিচের বাটন ব্যবহার করুন।", chat_id, reply_markup=get_main_keyboard(chat_id, chat_type))

    elif cmd == "/cancel":
        support_sessions.discard(uid)
        for d in [submission_sessions, withdraw_sessions, deposit_sessions, rps_sessions,
                  admin_add_mother_session, admin_add_mother_bulk_session, admin_approve_sessions]:
            d.pop(uid, None)
        session_activity.pop(uid, None)
        send_telegram_message("❌ চলমান প্রক্রিয়া বাতিল করা হয়েছে।", chat_id, reply_markup=get_main_keyboard(chat_id, chat_type))

    elif cmd == "/profile":
        show_profile(uid)

    elif cmd == "/balance":
        main = user_balances.get(uid, 0)
        game = game_balances.get(uid, 0)
        msg_text = f"💰 মূল ব্যালেন্স: {main} টাকা"
        if game > 0:
            msg_text += f"\n🎮 গেম ব্যালেন্স (মাদার কেনার জন্য): {game} টাকা"
        send_telegram_message(msg_text, uid)

    elif cmd == "/history":
        user_txns = [t for t in transactions if t["user_id"] == uid]
        if not user_txns:
            send_telegram_message("আপনার কোনো ট্রানজেকশন ইতিহাস নেই।", uid)
        else:
            recent = user_txns[-10:]
            lines = ["📜 **সাম্প্রতিক ট্রানজেকশন:**\n"]
            for t in reversed(recent):
                sign = "+" if t["amount"] >= 0 else ""
                date_str = datetime.datetime.fromtimestamp(t["timestamp"]).strftime("%d/%m/%Y %H:%M")
                lines.append(f"`{date_str}` | {t['description']} | {sign}{t['amount']} টাকা | ব্যালেন্স: {t['balance_after']} টাকা")
            send_telegram_message("\n".join(lines), uid, parse_mode="Markdown")

    # --- Admin commands (only for ADMIN_CHAT_ID) ---
    elif uid != ADMIN_CHAT_ID:
        send_telegram_message("❌ অজানা কমান্ড।", uid)
        return

    # Admin commands below
    elif cmd == "/addmother":
        start_add_mother_stock(uid)
    elif cmd == "/addbulkmother":
        start_add_mother_bulk(uid)
    elif cmd == "/motherstock":
        show_mother_stock_detail(uid)
    elif cmd == "/freemotherlist":
        show_free_mother_list(uid)
    elif cmd == "/approvewithdraw":
        if len(parts) < 2: send_telegram_message("Usage: /approvewithdraw <id>", ADMIN_CHAT_ID); return
        w_id = parts[1]
        with data_lock:
            for w in withdraw_requests:
                if w["id"] == w_id and w["status"] == "pending":
                    user_id = str(w["user_id"])
                    amount = w["amount"]
                    # Balance already deducted; just mark approved
                    w["status"] = "approved"
                    save_all()
                    send_telegram_message(f"✅ উইথড্র {w_id} অনুমোদিত হয়েছে।", ADMIN_CHAT_ID)
                    send_telegram_message(f"✅ আপনার {amount} টাকা উইথড্র অনুমোদিত হয়েছে।", user_id)
                    return
        send_telegram_message("❌ উইথড্র রিকোয়েস্ট পাওয়া যায়নি।", ADMIN_CHAT_ID)
    elif cmd == "/rejectwithdraw":
        if len(parts) < 2: send_telegram_message("Usage: /rejectwithdraw <id>", ADMIN_CHAT_ID); return
        w_id = parts[1]
        with data_lock:
            for w in withdraw_requests:
                if w["id"] == w_id and w["status"] == "pending":
                    user_id = str(w["user_id"])
                    amount = w["amount"]
                    w["status"] = "rejected"
                    # Refund the deducted amount
                    user_balances[user_id] = user_balances.get(user_id, 0) + amount
                    record_transaction(user_id, "withdraw_refund", amount, f"Rejected withdraw {w_id}")
                    save_all()
                    send_telegram_message(f"✅ উইথড্র {w_id} প্রত্যাখ্যান করে টাকা ফেরত দেওয়া হয়েছে।", ADMIN_CHAT_ID)
                    send_telegram_message(f"❌ আপনার {amount} টাকা উইথড্র রিকোয়েস্ট প্রত্যাখ্যান হয়েছে। টাকা ফেরত পেয়েছেন।", user_id)
                    return
        send_telegram_message("❌ উইথড্র রিকোয়েস্ট পাওয়া যায়নি।", ADMIN_CHAT_ID)
    elif cmd == "/approvedeposit":
        if len(parts) < 2: send_telegram_message("Usage: /approvedeposit <id>", ADMIN_CHAT_ID); return
        dep_id = parts[1]
        with data_lock:
            for d in deposit_requests:
                if d["id"] == dep_id and d["status"] == "pending":
                    user_id = str(d["user_id"])
                    user_balances[user_id] = user_balances.get(user_id, 0) + d["amount"]
                    d["status"] = "approved"
                    record_transaction(user_id, "deposit", d["amount"], f"Deposit {dep_id} ({d['method'].upper()})")
                    save_all()
                    send_telegram_message(f"✅ ডিপোজিট {dep_id} অনুমোদিত হয়েছে।", ADMIN_CHAT_ID)
                    send_telegram_message(f"✅ আপনার {d['amount']} টাকা ডিপোজিট অনুমোদিত হয়েছে।", user_id)
                    return
        send_telegram_message("❌ ডিপোজিট রিকোয়েস্ট পাওয়া যায়নি।", ADMIN_CHAT_ID)
    elif cmd == "/rejectdeposit":
        if len(parts) < 2: send_telegram_message("Usage: /rejectdeposit <id>", ADMIN_CHAT_ID); return
        dep_id = parts[1]
        with data_lock:
            for d in deposit_requests:
                if d["id"] == dep_id and d["status"] == "pending":
                    d["status"] = "rejected"
                    save_all()
                    send_telegram_message(f"✅ ডিপোজিট {dep_id} প্রত্যাখ্যান করা হয়েছে।", ADMIN_CHAT_ID)
                    send_telegram_message(f"❌ আপনার {d['amount']} টাকা ডিপোজিট রিকোয়েস্ট প্রত্যাখ্যান হয়েছে।", str(d["user_id"]))
                    return
        send_telegram_message("❌ ডিপোজিট রিকোয়েস্ট পাওয়া যায়নি।", ADMIN_CHAT_ID)
    elif cmd == "/setprice":
        if len(parts) < 3: send_telegram_message("Usage: /setprice <2fa/cookies> <amount>", ADMIN_CHAT_ID); return
        item, price = parts[1], float(parts[2])
        if price <= 0: send_telegram_message("❌ মূল্য ধনাত্মক হতে হবে।", ADMIN_CHAT_ID); return
        if item == "2fa": config["price_2fa"] = price
        elif item == "cookies": config["price_cookies"] = price
        else: send_telegram_message("আইটেম: 2fa বা cookies", ADMIN_CHAT_ID); return
        save_all()
        send_telegram_message(f"✅ {item} মূল্য আপডেট হয়েছে: {price} টাকা", ADMIN_CHAT_ID)
    elif cmd == "/setmotherprice":
        if len(parts) < 2: send_telegram_message("Usage: /setmotherprice <amount>", ADMIN_CHAT_ID); return
        price = float(parts[1])
        if price <= 0: send_telegram_message("❌ ধনাত্মক সংখ্যা দিন।", ADMIN_CHAT_ID); return
        config["mother_price"] = price
        save_all()
        send_telegram_message(f"✅ মাদার একাউন্ট মূল্য: {config['mother_price']} টাকা", ADMIN_CHAT_ID)
    elif cmd == "/setreferral":
        if len(parts) < 3: send_telegram_message("Usage: /setreferral <level1/level2> <percent>", ADMIN_CHAT_ID); return
        level, percent = parts[1], float(parts[2])
        if percent < 0 or percent > 100: send_telegram_message("❌ শতকরা ০-১০০ এর মধ্যে হতে হবে।", ADMIN_CHAT_ID); return
        if level == "level1": config["referral_level1"] = percent
        elif level == "level2": config["referral_level2"] = percent
        else: send_telegram_message("level1 বা level2", ADMIN_CHAT_ID); return
        save_all()
        send_telegram_message(f"✅ রেফারেল {level}: {percent}%", ADMIN_CHAT_ID)
    elif cmd == "/setbkash":
        if len(parts) < 2: send_telegram_message("Usage: /setbkash <number>", ADMIN_CHAT_ID); return
        config["bkash_number"] = parts[1]
        save_all()
        send_telegram_message(f"✅ বিকাশ নম্বর: {config['bkash_number']}", ADMIN_CHAT_ID)
    elif cmd == "/setnagad":
        if len(parts) < 2: send_telegram_message("Usage: /setnagad <number>", ADMIN_CHAT_ID); return
        config["nagad_number"] = parts[1]
        save_all()
        send_telegram_message(f"✅ নগদ নম্বর: {config['nagad_number']}", ADMIN_CHAT_ID)
    elif cmd == "/setchannel":
        if len(parts) < 2: send_telegram_message("Usage: /setchannel <channel_id>", ADMIN_CHAT_ID); return
        config["channel_id"] = parts[1]
        save_all()
        send_telegram_message(f"✅ চ্যানেল আইডি: {config['channel_id']}", ADMIN_CHAT_ID)
    elif cmd == "/maintenance":
        maintenance_mode = not maintenance_mode
        config["maintenance_mode"] = maintenance_mode
        save_all()
        send_telegram_message(f"🔧 মেইনটেনেন্স মোড {'চালু' if maintenance_mode else 'বন্ধ'}", ADMIN_CHAT_ID)
    elif cmd == "/broadcast":
        admin_broadcast_prompt(ADMIN_CHAT_ID)
    elif cmd == "/backup":
        save_data_to_channel()
        send_telegram_message("📁 ব্যাকআপ চ্যানেলে পাঠানো হয়েছে।", ADMIN_CHAT_ID)
    elif cmd == "/restore":
        if not msg.get("reply_to_message"):
            send_telegram_message("❌ দয়া করে ব্যাকআপ ফাইল বা ইনডেক্স মেসেজে রিপ্লাই দিয়ে /restore দিন।", ADMIN_CHAT_ID)
            return
        reply_msg = msg["reply_to_message"]
        try:
            session = get_bot_session()
            compressed = None
            if "document" in reply_msg:
                file_id = reply_msg["document"]["file_id"]
                file_info = session.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}", timeout=20).json()
                if not file_info.get("ok"):
                    send_telegram_message("❌ ফাইল ডাউনলোড করা যায়নি।", ADMIN_CHAT_ID)
                    return
                file_path = file_info["result"]["file_path"]
                compressed = session.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}", timeout=60).content
            elif "text" in reply_msg:
                text_content = reply_msg["text"]
                start = text_content.find('{')
                if start == -1:
                    send_telegram_message("❌ ইনডেক্স JSON খুঁজে পাওয়া যায়নি।", ADMIN_CHAT_ID)
                    return
                index = json.loads(text_content[start:])
                file_ids = index.get("file_ids")
                if file_ids:
                    combined = bytearray()
                    for fid in file_ids:
                        file_info = session.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={fid}", timeout=20).json()
                        if not file_info.get("ok"):
                            send_telegram_message(f"❌ পার্ট ফাইল ডাউনলোড ব্যর্থ।", ADMIN_CHAT_ID)
                            return
                        file_path = file_info["result"]["file_path"]
                        part_data = session.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}", timeout=60).content
                        combined.extend(part_data)
                    compressed = bytes(combined)
                else:
                    part_ids = index.get("parts", [])
                    if not part_ids:
                        send_telegram_message("❌ ইনডেক্সে কোনো পার্ট নেই।", ADMIN_CHAT_ID)
                        return
                    combined = bytearray()
                    for pid in part_ids:
                        part_msg = session.get(
                            f"https://api.telegram.org/bot{BOT_TOKEN}/getChatMessage?chat_id={ADMIN_CHAT_ID}&message_id={pid}",
                            timeout=20
                        ).json()
                        if not part_msg.get("ok") or "document" not in part_msg.get("result", {}):
                            send_telegram_message(f"❌ পার্ট মেসেজ {pid} পাওয়া যায়নি।", ADMIN_CHAT_ID)
                            return
                        file_id = part_msg["result"]["document"]["file_id"]
                        file_info = session.get(
                            f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}",
                            timeout=20
                        ).json()
                        if not file_info.get("ok"):
                            send_telegram_message(f"❌ পার্ট {pid} ডাউনলোড ব্যর্থ।", ADMIN_CHAT_ID)
                            return
                        file_path = file_info["result"]["file_path"]
                        part_data = session.get(
                            f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}",
                            timeout=60
                        ).content
                        combined.extend(part_data)
                    compressed = bytes(combined)
            else:
                send_telegram_message("❌ রিপ্লাই করা মেসেজে কোনো ডকুমেন্ট বা টেক্সট নেই।", ADMIN_CHAT_ID)
                return

            data = restore_data_from_payload(compressed)
            apply_restored_data(data)
            send_telegram_message("✅ ডেটা সফলভাবে রিস্টোর হয়েছে!", ADMIN_CHAT_ID)
        except Exception as e:
            logger.exception("Manual restore error:")
            send_telegram_message(f"❌ রিস্টোর ব্যর্থ: {e}", ADMIN_CHAT_ID)
    elif cmd == "/userlist":
        with data_lock:
            users = list(subscribed_users)
        send_telegram_message(f"মোট সাবস্ক্রাইবার: {len(users)}", ADMIN_CHAT_ID)
    elif cmd == "/usermessage":
        if len(parts) < 3: send_telegram_message("Usage: /usermessage <user_id> <message>", ADMIN_CHAT_ID); return
        target = parts[1]
        message = " ".join(parts[2:])
        send_telegram_message(f"📨 অ্যাডমিনের মেসেজ:\n{message}", target)
        send_telegram_message(f"✅ {target} কে মেসেজ পাঠানো হয়েছে।", ADMIN_CHAT_ID)
    elif cmd == "/adminprofile":
        if len(parts) < 2: send_telegram_message("Usage: /adminprofile <user_id>", ADMIN_CHAT_ID); return
        show_profile(parts[1])
    else:
        send_telegram_message("❌ অজানা অ্যাডমিন কমান্ড।", ADMIN_CHAT_ID)

# ================== MAIN UPDATE HANDLER ==================
def handle_telegram_commands():
    global last_update_id, user_versions, maintenance_mode
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/getUpdates"
    while True:
        try:
            params = {"timeout": 30, "allowed_updates": ["message", "callback_query", "inline_query"]}
            if last_update_id:
                params["offset"] = last_update_id + 1
            session = get_bot_session()
            resp = session.get(url, params=params, timeout=35).json()
            if resp.get("ok") and resp.get("result"):
                for update in resp["result"]:
                    last_update_id = update["update_id"]

                    if "inline_query" in update:
                        handle_inline_query(update["inline_query"])
                        continue

                    if "callback_query" in update:
                        cb = update["callback_query"]
                        chat_id = str(cb["message"]["chat"]["id"])
                        data = cb["data"]
                        from_user = cb.get("from", {})
                        user_info[chat_id] = from_user.get("username") or from_user.get("first_name", f"ID:{chat_id}")
                        chat_type = cb["message"]["chat"]["type"]
                        answer_callback_query(cb["id"])

                        # Removed version check here to avoid interrupting active sessions

                        if data == "cancel_broadcast" and chat_id == ADMIN_CHAT_ID:
                            broadcast_sessions.pop(ADMIN_CHAT_ID, None)
                            send_telegram_message("❌ ব্রডকাস্ট বাতিল।", chat_id, reply_markup=admin_panel_keyboard())
                            continue
                        if data == "cancel_session":
                            for d in [admin_add_mother_session, admin_add_mother_bulk_session,
                                      admin_approve_sessions, broadcast_sessions,
                                      submission_sessions, withdraw_sessions, deposit_sessions, rps_sessions]:
                                d.pop(chat_id, None)
                            support_sessions.discard(chat_id)
                            session_activity.pop(chat_id, None)
                            kb = admin_panel_keyboard() if chat_id == ADMIN_CHAT_ID else get_main_keyboard(chat_id, chat_type)
                            send_telegram_message("❌ প্রক্রিয়া বাতিল করা হয়েছে।", chat_id, reply_markup=kb)
                            continue

                        # RPS
                        if data.startswith("rps_"):
                            process_rps_callback(chat_id, data[4:])
                            continue
                        # Other callbacks
                        if data == "sub_cookies":
                            start_submission(chat_id, "cookies") if not config.get("lock_cookies") else send_telegram_message("🔒 বন্ধ", chat_id)
                        elif data == "sub_2fa":
                            start_submission(chat_id, "2fa") if not config.get("lock_2fa") else send_telegram_message("🔒 বন্ধ", chat_id)
                        elif data in ["lock_2fa","lock_cookies"] and chat_id == ADMIN_CHAT_ID:
                            key = "lock_2fa" if data == "lock_2fa" else "lock_cookies"
                            config[key] = not config.get(key, False)
                            save_all()
                            send_telegram_message(f"{'2FA' if key=='lock_2fa' else 'কুকিজ'} {'🔒 বন্ধ' if config[key] else '🔓 চালু'}", chat_id)
                        elif data.startswith("getfile_") and chat_id == ADMIN_CHAT_ID:
                            sub = next((s for s in submissions if s["id"] == data[8:]), None)
                            if sub and "file_id" in sub:
                                session = get_bot_session()
                                session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                                             json={"chat_id": ADMIN_CHAT_ID, "document": sub["file_id"]})
                        elif data.startswith("approve_") and chat_id == ADMIN_CHAT_ID:
                            admin_approve_start(data[8:])
                        elif data.startswith("reject_") and chat_id == ADMIN_CHAT_ID:
                            sub_id = data[7:]
                            if reject_submission(sub_id):
                                send_telegram_message(f"✅ সাবমিশন {sub_id} রিজেক্ট করা হয়েছে।", ADMIN_CHAT_ID)
                            else:
                                send_telegram_message("❌ সাবমিশন পাওয়া যায়নি।", ADMIN_CHAT_ID)
                        elif data == "set_target":
                            submission_sessions[chat_id] = {"step": "target_amount"}
                            session_activity[chat_id] = time.time()
                            send_telegram_message("🎯 মাসিক টার্গেট কত টাকা?", chat_id,
                                                 reply_markup={"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]})
                        elif data.startswith("bc_") and chat_id == ADMIN_CHAT_ID:
                            broadcast_sessions[ADMIN_CHAT_ID] = {"type": data[3:]}
                            session_activity[ADMIN_CHAT_ID] = time.time()
                            prompts = {"text":"টেক্সট লিখুন","photo":"ছবি পাঠান","document":"ফাইল পাঠান","voice":"ভয়েস পাঠান"}
                            send_telegram_message(f"📢 {prompts.get(data[3:], '')} (সবাইকে পাঠানো হবে):", ADMIN_CHAT_ID)
                        elif data.startswith("delmotherstock_") and chat_id == ADMIN_CHAT_ID:
                            idx = int(data.split("_")[1])
                            with data_lock:
                                if 0 <= idx < len(mother_stock):
                                    deleted = mother_stock.pop(idx)
                                    schedule_save()
                                    show_mother_stock_detail_refresh(chat_id, cb["message"]["message_id"])
                                    send_telegram_message(f"🗑️ মাদার স্টক থেকে {deleted['username']} মুছে ফেলা হয়েছে।", ADMIN_CHAT_ID)
                        elif data.startswith("delfreemother_") and chat_id == ADMIN_CHAT_ID:
                            idx = int(data.split("_")[1])
                            with data_lock:
                                if 0 <= idx < len(mother_accounts):
                                    deleted = mother_accounts.pop(idx)
                                    schedule_save()
                                    show_free_mother_list_refresh(chat_id, cb["message"]["message_id"])
                                    send_telegram_message(f"🗑️ ফ্রি মাদার {deleted['username']} মুছে ফেলা হয়েছে।", ADMIN_CHAT_ID)
                        elif data.startswith("motherstock_page_") and chat_id == ADMIN_CHAT_ID:
                            page = int(data.split("_")[2])
                            show_mother_stock_detail(chat_id, page=page, message_id=cb["message"]["message_id"])
                        elif data == "close_motherstock" and chat_id == ADMIN_CHAT_ID:
                            send_telegram_message("📦 মাদার স্টক তালিকা বন্ধ।", chat_id, reply_markup=admin_panel_keyboard())
                        elif data.startswith("freemother_page_") and chat_id == ADMIN_CHAT_ID:
                            page = int(data.split("_")[2])
                            show_free_mother_list(chat_id, page=page, message_id=cb["message"]["message_id"])
                        elif data == "close_freemotherlist" and chat_id == ADMIN_CHAT_ID:
                            send_telegram_message("🎁 ফ্রি মাদার তালিকা বন্ধ।", chat_id, reply_markup=admin_panel_keyboard())
                        elif data.startswith("admin_profile_") and chat_id == ADMIN_CHAT_ID:
                            target = data[14:]
                            send_telegram_message(get_profile_text(target), chat_id)
                        elif data.startswith("depmethod_"):
                            method = data[10:]
                            deposit_sessions[chat_id] = {"step": "amount", "method": method}
                            session_activity[chat_id] = time.time()
                            number = config.get(f"{method}_number", "সেট করা হয়নি")
                            cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
                            send_telegram_message(
                                f"আপনার {method.upper()} নম্বর থেকে **{number}** নম্বরে টাকা পাঠিয়ে নিচে ট্রানজেকশন আইডি দিন।\n\n"
                                "প্রথমে কত টাকা পাঠিয়েছেন তা লিখুন (শুধু সংখ্যা):\nবাতিল করতে /cancel",
                                chat_id, reply_markup=cancel_kb)
                        elif data.startswith("withmethod_"):
                            method = data[11:]
                            withdraw_sessions[chat_id] = {"step": "amount", "method": method}
                            session_activity[chat_id] = time.time()
                            cancel_kb = {"inline_keyboard": [[{"text": "❌ বাতিল", "callback_data": "cancel_session"}]]}
                            send_telegram_message("💸 কত টাকা উইথড্র করতে চান? (শুধু সংখ্যা)\nবাতিল করতে /cancel", chat_id, reply_markup=cancel_kb)
                        elif data == "show_history":
                            user_txns = [t for t in transactions if t["user_id"] == chat_id]
                            if not user_txns:
                                send_telegram_message("ট্রানজেকশন ইতিহাস নেই।", chat_id)
                            else:
                                recent = user_txns[-10:]
                                lines = ["📜 **সাম্প্রতিক ট্রানজেকশন:**\n"]
                                for t in reversed(recent):
                                    sign = "+" if t["amount"] >= 0 else ""
                                    date_str = datetime.datetime.fromtimestamp(t["timestamp"]).strftime("%d/%m/%Y %H:%M")
                                    lines.append(f"`{date_str}` | {t['description']} | {sign}{t['amount']} টাকা | ব্যালেন্স: {t['balance_after']} টাকা")
                                send_telegram_message("\n".join(lines), chat_id, parse_mode="Markdown")
                        continue

                    if "message" in update:
                        msg = update["message"]
                        chat_id = str(msg["chat"]["id"])
                        chat_type = msg["chat"]["type"]
                        text = msg.get("text", "").strip()
                        from_user = msg.get("from", {})
                        user_info[chat_id] = from_user.get("username") or from_user.get("first_name", f"ID:{chat_id}")

                        if chat_type == "private":
                            current_version = config.get("bot_version", "1.0")
                            if user_versions.get(chat_id, "0") != current_version:
                                # Force restart only via message, not callbacks
                                for d in [submission_sessions, withdraw_sessions, deposit_sessions,
                                          admin_add_mother_session, admin_add_mother_bulk_session,
                                          admin_approve_sessions, broadcast_sessions, rps_sessions]:
                                    d.pop(chat_id, None)
                                support_sessions.discard(chat_id)
                                session_activity.pop(chat_id, None)
                                send_telegram_message("🔄 বট আপডেট হয়েছে! /start দিন।", chat_id, reply_markup=get_main_keyboard(chat_id, chat_type))
                                user_versions[chat_id] = current_version
                                schedule_save()
                                continue

                        if maintenance_mode and chat_id != ADMIN_CHAT_ID:
                            send_telegram_message("🔧 রক্ষণাবেক্ষণ মোড চালু আছে।", chat_id)
                            continue
                        if chat_id in support_sessions and text.lower() in ["/cancel", "/start"]:
                            support_sessions.discard(chat_id)
                            session_activity.pop(chat_id, None)
                            send_telegram_message("❌ সাপোর্ট বন্ধ।", chat_id, reply_markup=get_main_keyboard(chat_id, chat_type))
                            continue
                        if chat_id in support_sessions:
                            forward_support_message(chat_id, msg)
                            continue
                        if chat_id == ADMIN_CHAT_ID and ADMIN_CHAT_ID in broadcast_sessions and text.lower() in ["/cancel", "/start"]:
                            del broadcast_sessions[ADMIN_CHAT_ID]
                            session_activity.pop(ADMIN_CHAT_ID, None)
                            send_telegram_message("❌ ব্রডকাস্ট বাতিল।", ADMIN_CHAT_ID, reply_markup=admin_panel_keyboard())
                            continue
                        # Active sessions
                        if chat_id == ADMIN_CHAT_ID and ADMIN_CHAT_ID in admin_approve_sessions:
                            process_admin_approve_step(chat_id, text)
                            continue
                        if chat_id in deposit_sessions:
                            process_deposit_step(chat_id, text)
                            continue
                        if chat_id in submission_sessions:
                            session = submission_sessions[chat_id]
                            if session.get("step") == "target_amount":
                                if text.lower() in ["/cancel","/start"]:
                                    del submission_sessions[chat_id]
                                    session_activity.pop(chat_id, None)
                                    send_telegram_message("❌ বাতিল।", chat_id, reply_markup=get_main_keyboard(chat_id, chat_type))
                                    continue
                                try:
                                    target = float(text.strip())
                                    if target < 0: raise ValueError
                                except:
                                    send_telegram_message("⚠️ সঠিক সংখ্যা দিন।", chat_id); continue
                                with data_lock:
                                    init_leaderboard_entry(chat_id)
                                    leaderboard[str(chat_id)]["monthly_target"] = target
                                    schedule_save()
                                del submission_sessions[chat_id]
                                session_activity.pop(chat_id, None)
                                send_telegram_message(f"✅ টার্গেট {target} টাকা সেট হয়েছে।", chat_id, reply_markup=get_main_keyboard(chat_id, chat_type))
                                continue
                            elif session.get("type") == "mother_buy":
                                process_mother_buy_step(chat_id, text)
                                continue
                            else:
                                process_submission_step(chat_id, text, user_info[chat_id])
                                continue
                        if chat_id in admin_add_mother_session:
                            process_add_mother_step(chat_id, text)
                            continue
                        if chat_id in admin_add_mother_bulk_session:
                            process_add_mother_bulk_step(chat_id, text)
                            continue
                        if chat_id in withdraw_sessions:
                            process_withdraw_step(chat_id, text)
                            continue
                        # Broadcast content
                        if chat_id == ADMIN_CHAT_ID and ADMIN_CHAT_ID in broadcast_sessions:
                            bc_type = broadcast_sessions[ADMIN_CHAT_ID]["type"]
                            if bc_type == "text" and text:
                                send_telegram_message("📢 ব্রডকাস্ট শুরু...", ADMIN_CHAT_ID)
                                threading.Thread(target=broadcast_message, args=(text,), daemon=True).start()
                            elif bc_type == "photo" and "photo" in msg:
                                file_id = msg["photo"][-1]["file_id"]
                                caption = msg.get("caption", "")
                                send_telegram_message("📢 ব্রডকাস্ট শুরু...", ADMIN_CHAT_ID)
                                threading.Thread(target=broadcast_media, args=("photo", file_id, caption), daemon=True).start()
                            elif bc_type == "document" and "document" in msg:
                                file_id = msg["document"]["file_id"]
                                caption = msg.get("caption", "")
                                send_telegram_message("📢 ব্রডকাস্ট শুরু...", ADMIN_CHAT_ID)
                                threading.Thread(target=broadcast_media, args=("document", file_id, caption), daemon=True).start()
                            elif bc_type == "voice" and "voice" in msg:
                                file_id = msg["voice"]["file_id"]
                                send_telegram_message("📢 ব্রডকাস্ট শুরু...", ADMIN_CHAT_ID)
                                threading.Thread(target=broadcast_media, args=("voice", file_id, ""), daemon=True).start()
                            else:
                                send_telegram_message("❌ ভুল ফরম্যাট।", ADMIN_CHAT_ID)
                            del broadcast_sessions[ADMIN_CHAT_ID]
                            session_activity.pop(ADMIN_CHAT_ID, None)
                            continue
                        # Button shortcuts
                        if text == "💼 একাউন্ট সাবমিট":
                            send_telegram_message("কোন ধরণের একাউন্ট?", chat_id,
                                                 reply_markup={"inline_keyboard": [
                                                     [{"text": "🍪 কুকিজ", "callback_data": "sub_cookies"}],
                                                     [{"text": "🔐 2FA", "callback_data": "sub_2fa"}]
                                                 ]})
                        elif text == "👤 প্রোফাইল": show_profile(chat_id)
                        elif text == "👥 রেফারেল":
                            link = f"https://t.me/{BOT_USERNAME}?start=ref_{chat_id}"
                            send_telegram_message(f"🔗 রেফারেল লিংক:\n{link}", chat_id)
                        elif text == "💰 ব্যালেন্স":
                            main = user_balances.get(str(chat_id), 0)
                            game = game_balances.get(str(chat_id), 0)
                            m = f"💰 মূল ব্যালেন্স: {main} টাকা"
                            if game > 0: m += f"\n🎮 গেম ব্যালেন্স: {game} টাকা"
                            send_telegram_message(m, chat_id)
                        elif text == "💳 ডিপোজিট": start_deposit(chat_id)
                        elif text == "💸 উইথড্র": start_withdraw(chat_id)
                        elif text == "📊 লিডারবোর্ড": show_leaderboard(chat_id)
                        elif text == "🎁 ফ্রি মাদার একাউন্ট": handle_get_free_mother(chat_id)
                        elif text == "🛒 মাদার একাউন্ট কিনুন": start_buy_mother(chat_id)
                        elif text == "📞 সাপোর্ট": start_support(chat_id)
                        elif text == "🎮 RPS গেম": start_rps(chat_id)
                        elif text == "🛠️ অ্যাডমিন প্যানেল" and chat_id == ADMIN_CHAT_ID:
                            send_telegram_message("অ্যাডমিন প্যানেল", chat_id, reply_markup=admin_panel_keyboard())
                        elif text.startswith("/"):
                            handle_commands(chat_id, text, chat_type, msg)
                        elif chat_type == "private" and text:
                            send_telegram_message("❌ অজানা কমান্ড।", chat_id)
        except Exception as e:
            logger.exception("Main loop error:")
        time.sleep(1)

# ================== DAILY TASKS ==================
def daily_task_loop():
    global last_morning_sent_date, last_evening_sent_date
    while True:
        now = datetime.datetime.now()
        today_str = str(now.date())
        # Morning notification at 03:00
        if now.hour == 3 and last_morning_sent_date != today_str:
            last_morning_sent_date = today_str
            for uid in list(subscribed_users):
                progress = get_target_progress(uid)
                if progress:
                    msg = (
                        f"🌅 সুপ্রভাত!\nমাসিক টার্গেট: {progress['target']} টাকা\n"
                        f"চলতি মাসের আয়: {progress['current_income']} টাকা\n"
                        f"⏳ বাকি: {progress['remaining']} টাকা ({progress['days_left']} দিন)\n"
                        f"📌 আজকের গড় প্রয়োজন: {progress['daily_income_needed']:.1f} টাকা"
                    )
                    send_telegram_message(msg, uid)
                    time.sleep(0.05)
        # Evening notification at 15:00
        if now.hour == 15 and last_evening_sent_date != today_str:
            last_evening_sent_date = today_str
            for uid in list(subscribed_users):
                progress = get_target_progress(uid)
                if progress and progress['remaining'] > 0:
                    msg = (
                        f"⏰ শুভ সন্ধ্যা!\nবাকি: {progress['remaining']} টাকা\n"
                        f"আজকের আয়: {progress['today_income']} টাকা\n"
                        f"প্রয়োজনীয়: 2FA {progress['daily_2fa_needed']:.1f} টি, কুকিজ {progress['daily_cookies_needed']:.1f} টি"
                    )
                    send_telegram_message(msg, uid)
                    time.sleep(0.05)
        time.sleep(60)

# ================== DUPLICATE CLEANUP ==================
def duplicate_cleanup_loop():
    while True:
        time.sleep(172800)
        with data_lock:
            submitted_usernames.clear()
            save_all()
        logger.info("Duplicate username set cleared")

def user_versions_cleanup_loop():
    while True:
        time.sleep(604800)
        with data_lock:
            active = set(subscribed_users)
            for uid in list(user_versions):
                if uid not in active:
                    del user_versions[uid]
            save_all()
        logger.info("Cleaned old user_versions entries")

# ================== SESSION CLEANUP ==================
def session_cleanup_loop():
    while True:
        time.sleep(60)
        now = time.time()
        expired = []
        for cid, timestamp in list(session_activity.items()):
            if now - timestamp > 900:  # 15 minutes
                expired.append(cid)
        for cid in expired:
            for d in [submission_sessions, withdraw_sessions, deposit_sessions,
                      admin_add_mother_session, admin_add_mother_bulk_session,
                      admin_approve_sessions, rps_sessions, broadcast_sessions]:
                d.pop(cid, None)
            support_sessions.discard(cid)
            session_activity.pop(cid, None)
            logger.info(f"Cleaned up stale session for {cid}")

# ================== FLASK ==================
@app.route("/")
def home():
    return "Bot Running!"

# ================== MAIN ==================
if __name__ == "__main__":
    load_all()
    auto_restore_from_channel()
    new_build_id = os.environ.get("RENDER_GIT_COMMIT", uuid.uuid4().hex)
    old_build_id = config.get("build_id", "")
    if old_build_id != new_build_id:
        config["bot_version"] = str(int(time.time()))
        config["build_id"] = new_build_id
        save_all()
        logger.info(f"Auto version updated to {config['bot_version']}")
    try:
        session = get_bot_session()
        session.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteWebhook?drop_pending_updates=true")
    except: pass
    threading.Thread(target=auto_backup_loop, daemon=True).start()
    threading.Thread(target=daily_task_loop, daemon=True).start()
    threading.Thread(target=duplicate_cleanup_loop, daemon=True).start()
    threading.Thread(target=user_versions_cleanup_loop, daemon=True).start()
    threading.Thread(target=session_cleanup_loop, daemon=True).start()
    threading.Thread(target=handle_telegram_commands, daemon=True).start()
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
